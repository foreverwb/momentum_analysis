#!/usr/bin/env python3
"""
Momentum Radar 命令行工具

支持的命令:
- uploads: 上传 ETF Holdings 文件（支持 xlsx、xls、csv 格式）
- update: 更新 ETF Holdings 数据（与 uploads 相同，更语义化的命令）
- finviz etfs: 导入 Finviz JSON/CSV ETF 数据
- mc etfs: 导入 MarketChameleon ETF 数据（支持整文件或按覆盖范围导入）
- refresh etfs: 后台串行刷新多个 ETF
- refresh holdings: 后台串行刷新多个 ETF holdings
- init: 初始化数据库和默认数据

使用示例:
    # 上传板块 ETF holdings（日期可选，默认为当天）
    python -m app.cli uploads -t sector -a XLK holdings.xlsx
    python -m app.cli uploads -d 2026-01-25 -t sector -a XLK holdings.xlsx
    
    # 上传行业 ETF holdings（需要指定父板块）
    python -m app.cli uploads -d 2026-01-25 -t industry -s XLK -a SOXX holdings.csv
    
    # 更新数据（日期可选，默认为当天）
    python -m app.cli update -t sector -a XLE xle.xlsx
    python -m app.cli update -d 2026-01-28 -t sector -a XLE xle.xlsx

    # 导入 MarketChameleon ETF 数据（整文件）
    python -m app.cli mc etfs -f marketchameleon_etfs.json

    # 导入 MarketChameleon 数据（多个 ETF + weight 覆盖）
    python -m app.cli mc etfs -s "XLK,XLC,XLV" -w 85 -f marketchameleon.json

    # 导入 Finviz 数据（支持 JSON/CSV）
    python -m app.cli finviz etfs -f finviz_export.csv
    python -m app.cli finviz etfs -s "XLK,XLC,XLV" -w 85 -f finviz_export.csv

    # 后台提交 refresh 任务（命令立即返回）
    python -m app.cli refresh etfs -s "XLK,XLF,SOXX"
    python -m app.cli refresh holdings -s "XLK,SOXX" -w t-20

    # 初始化数据库
    python -m app.cli init
"""

import argparse
import importlib.util
import sys
import os
import json
import re
import time
from urllib import error as urllib_error
from urllib import request as urllib_request
from datetime import datetime, date, timezone
from pathlib import Path

MC_IMPORT_FIELDS = [
    'symbol', 'RelVolTo90D', 'CallVolume', 'PutVolume', 'Earnings',
    'PutPct', 'IV30', 'IVR', 'HV20', 'PriceChgPct', 'IV30ChgPct'
]
PROVIDER_COMMANDS = {"finviz", "mc"}
PROVIDER_ETF_SUBCOMMAND = "etfs"
_SHARE_CLASS_ALIAS_PATTERN = re.compile(r"^([A-Z][A-Z0-9]{0,5})[.-]([A-Z])$")
SQLITE_LOCK_RETRY_ATTEMPTS = 3
DEFAULT_API_BASE_URL = os.environ.get("MOMENTUM_API_BASE_URL", "http://127.0.0.1:8000")
REFRESH_JOB_TERMINAL_STATUSES = {"completed", "failed", "cancelled"}
REFRESH_POST_SUBMIT_POLL_SECONDS = 2.0
REFRESH_POST_SUBMIT_POLL_INTERVAL_SECONDS = 0.25

# 添加 backend 根目录到 Python 路径，兼容直接执行 app/cli.py。
BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_ROOT))

from app.core.time_utils import beijing_today, utc_now_naive


TOP_LEVEL_COMMANDS = {
    "uploads",
    "update",
    "init",
    "list-etfs",
    "list-holdings",
    "finviz",
    "mc",
    "refresh",
}
FILE_PREFIX_ATTR_BY_COMMAND = {
    "uploads": "holdings_file_prefix",
    "update": "holdings_file_prefix",
    "finviz": "finviz_file_prefix",
    "mc": "mc_file_prefix",
}


def _running_in_venv() -> bool:
    base_prefix = getattr(sys, "base_prefix", sys.prefix)
    return base_prefix != sys.prefix


def _maybe_reexec_in_venv() -> None:
    if os.environ.get("CLI_NO_VENV") == "1":
        return
    if _running_in_venv():
        return
    backend_dir = Path(__file__).resolve().parents[1]
    project_root = backend_dir.parent
    venv_dirs = [backend_dir / ".venv", project_root / ".venv"]
    candidates = []
    if os.name == "nt":
        candidates = [venv / "Scripts" / "python.exe" for venv in venv_dirs]
    else:
        for venv in venv_dirs:
            candidates.extend([venv / "bin" / "python3", venv / "bin" / "python"])
    for candidate in candidates:
        if candidate.exists():
            os.execv(str(candidate), [str(candidate), "-m", "app.cli", *sys.argv[1:]])


def _ensure_dependency(module_name: str) -> None:
    if importlib.util.find_spec(module_name) is not None:
        return
    print(f"\n错误: 缺少依赖 {module_name}")
    print("请先安装后端依赖，再运行命令。推荐方式:")
    print("  cd backend")
    print("  python3 -m venv .venv")
    print("  source .venv/bin/activate  # Windows: .venv\\Scripts\\activate")
    print("  python -m pip install -r requirements.txt")
    print("\n也可以直接使用虚拟环境的 Python 运行:")
    print("  cd backend")
    print("  .venv/bin/python -m app.cli <命令> ...")
    sys.exit(1)


def _ensure_db_dependencies() -> None:
    _ensure_dependency("sqlalchemy")


_maybe_reexec_in_venv()


def _infer_cli_invoked_command(prog: str | None = None) -> str | None:
    program = Path(prog or sys.argv[0]).name
    candidates = [program, Path(program).stem]
    stem = Path(program).stem
    if stem.endswith("-script"):
        candidates.append(stem[:-7])
    for candidate in candidates:
        if candidate in TOP_LEVEL_COMMANDS:
            return candidate
    return None


def _is_plain_filename(raw_value: str) -> bool:
    if not raw_value:
        return False
    if raw_value.startswith("."):
        return False
    separators = [os.sep]
    if os.altsep:
        separators.append(os.altsep)
    if any(separator in raw_value for separator in separators):
        return False
    return True


def _resolve_cli_downloads_dir(config) -> Path:
    configured = (config.cli.downloads_dir or "").strip()
    downloads_dir = Path(configured).expanduser() if configured else Path.home() / "Downloads"
    if downloads_dir.is_absolute():
        return downloads_dir.resolve()

    if config.loaded_from_file and config.config_path:
        base_dir = Path(config.config_path).expanduser().resolve().parent
    else:
        base_dir = BACKEND_ROOT.parent
    return (base_dir / downloads_dir).resolve()


def _resolve_cli_file_prefix(command: str, config) -> str:
    command_prefix_attr = FILE_PREFIX_ATTR_BY_COMMAND.get(command)
    if command_prefix_attr:
        command_prefix = getattr(config.cli, command_prefix_attr, "").strip()
        if command_prefix:
            return command_prefix
    return (config.cli.file_prefix or "").strip()


def resolve_cli_file_arg(file_value: str, command: str) -> str:
    raw_value = str(file_value).strip()
    if not raw_value:
        return raw_value

    candidate = Path(raw_value).expanduser()
    if candidate.is_absolute():
        return str(candidate)

    direct_candidate = (Path.cwd() / candidate).resolve()
    if candidate.exists():
        return str(candidate.resolve())
    if direct_candidate.exists():
        return str(direct_candidate)
    if not _is_plain_filename(raw_value):
        return str(direct_candidate)

    try:
        from app.core.broker_config import load_broker_config
    except Exception:
        return raw_value

    config = load_broker_config()
    downloads_dir = _resolve_cli_downloads_dir(config)
    file_prefix = _resolve_cli_file_prefix(command, config)

    prefixed_candidate = None
    if file_prefix and not raw_value.startswith(file_prefix):
        prefixed_candidate = downloads_dir / f"{file_prefix}{raw_value}"
        if prefixed_candidate.exists():
            return str(prefixed_candidate.resolve())

    downloads_candidate = downloads_dir / raw_value
    if downloads_candidate.exists():
        return str(downloads_candidate.resolve())

    return raw_value


def normalize_cli_argv(argv=None, prog: str | None = None) -> list:
    raw_args = list(sys.argv[1:] if argv is None else argv)
    invoked_command = _infer_cli_invoked_command(prog)
    if invoked_command and (not raw_args or raw_args[0] not in TOP_LEVEL_COMMANDS):
        raw_args = [invoked_command, *raw_args]

    if not raw_args:
        return raw_args

    command = raw_args[0]
    if command not in PROVIDER_COMMANDS:
        return raw_args

    if len(raw_args) >= 2 and raw_args[1] == PROVIDER_ETF_SUBCOMMAND:
        return raw_args

    if len(raw_args) == 1 or raw_args[1].startswith('-'):
        return [command, PROVIDER_ETF_SUBCOMMAND, *raw_args[1:]]

    return raw_args


def _finalize_cli_args(parser: argparse.ArgumentParser, args):
    if args.command in {"uploads", "update"}:
        positional_file = getattr(args, "file_arg", None)
        optional_file = getattr(args, "file", None)
        if positional_file and optional_file:
            parser.error("uploads/update 不能同时使用位置参数 file 和 -f/--file")
        resolved_file = optional_file or positional_file
        if not resolved_file:
            parser.error("uploads/update 需要提供文件路径，可使用位置参数 file 或 -f/--file")
        args.file = resolved_file

    if args.command in FILE_PREFIX_ATTR_BY_COMMAND and getattr(args, "file", None):
        args.file = resolve_cli_file_arg(args.file, args.command)

    return args


def parse_cli_args(argv=None, prog: str | None = None):
    parser = build_parser()
    args = parser.parse_args(normalize_cli_argv(argv, prog=prog))
    return _finalize_cli_args(parser, args)


def _today_iso_date() -> str:
    return beijing_today().isoformat()


def parse_xlsx_holdings(file_path: str) -> list:
    """解析 xlsx 文件，提取 Ticker 和 Weight 列"""
    try:
        import openpyxl
        import re
        
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # 在 read_only 模式下需要重置维度以正确读取所有列
        sheet.reset_dimensions()
        
        # 获取表头 - 使用 iter_rows 确保在 read_only 模式下正确读取所有列
        headers = []
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).strip() if cell else "" for cell in row]
            break
        
        def normalize_header(value) -> str:
            if value is None:
                return ""
            text = str(value).strip().lower().replace("\u00a0", " ")
            text = re.sub(r"\s+", " ", text)
            return re.sub(r"[^a-z0-9]", "", text)

        # 查找 Ticker 和 Weight 列索引
        ticker_idx = None
        weight_idx = None
        
        for idx, header in enumerate(headers):
            header_key = normalize_header(header)
            if header_key in {"ticker", "symbol"}:
                ticker_idx = idx
            elif header_key == "weight":
                weight_idx = idx
        
        if ticker_idx is None:
            raise ValueError("未找到 'Ticker' 列")
        if weight_idx is None:
            raise ValueError("未找到 'Weight' 列")
        
        # 解析数据行
        holdings = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) > max(ticker_idx, weight_idx):
                ticker = row[ticker_idx]
                weight = row[weight_idx]
                
                if ticker and weight is not None:
                    holdings.append({
                        "row": row_idx,
                        "ticker": str(ticker).strip(),
                        "weight": weight
                    })
        
        return holdings
        
    except ImportError:
        print("错误: 需要安装 openpyxl 库")
        print("请运行: pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"错误: 解析 xlsx 文件失败 - {e}")
        sys.exit(1)


def parse_csv_holdings(file_path: str) -> list:
    """解析 csv 文件，提取 Ticker 和 Weight 列"""
    try:
        import csv
        import re
        
        holdings = []
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            def normalize_header(value) -> str:
                if value is None:
                    return ""
                text = str(value).strip().lower().replace("\u00a0", " ")
                text = re.sub(r"\s+", " ", text)
                return re.sub(r"[^a-z0-9]", "", text)

            # 查找 Ticker 和 Weight 列（支持 Weight (%)/Weight %）
            normalized = {normalize_header(name): name for name in reader.fieldnames} if reader.fieldnames else {}
            
            ticker_col = normalized.get('ticker') or normalized.get('symbol')
            weight_col = normalized.get('weight')
            
            if not ticker_col:
                raise ValueError("未找到 'Ticker' 列")
            if not weight_col:
                raise ValueError("未找到 'Weight' 列")
            
            for row_idx, row in enumerate(reader, start=2):
                ticker = row.get(ticker_col)
                weight = row.get(weight_col)
                
                if ticker and weight is not None:
                    # 处理可能带有千分位逗号的数字
                    if isinstance(weight, str):
                        weight = weight.replace(',', '').replace('%', '')
                    holdings.append({
                        "row": row_idx,
                        "ticker": str(ticker).strip(),
                        "weight": weight
                    })
        
        return holdings
        
    except Exception as e:
        print(f"错误: 解析 csv 文件失败 - {e}")
        sys.exit(1)


def validate_holdings(holdings: list) -> tuple:
    """验证并过滤持仓数据"""
    import re

    def normalize_weight(value):
        if isinstance(value, str):
            cleaned = value.strip().replace(",", "")
            if cleaned.endswith("%"):
                cleaned = cleaned[:-1].strip()
            return cleaned
        return value
    
    def is_valid_ticker(ticker: str) -> bool:
        if not ticker or not isinstance(ticker, str):
            return False
        ticker = ticker.strip()
        if not ticker:
            return False
        pattern = r'^[A-Za-z][A-Za-z0-9.\-]*$'
        return bool(re.match(pattern, ticker))
    
    valid_holdings = []
    skipped = []
    
    for h in holdings:
        ticker = h.get("ticker", "")
        weight = h.get("weight")
        row = h.get("row", "unknown")
        
        # 验证 Ticker
        if not is_valid_ticker(ticker):
            skipped.append({
                "row": str(row),
                "ticker": ticker,
                "reason": "Ticker 为空或不是有效的英文字符"
            })
            continue
        
        # 验证 Weight
        try:
            weight_float = float(normalize_weight(weight))
            if weight_float <= 0:
                skipped.append({
                    "row": str(row),
                    "ticker": ticker,
                    "reason": f"Weight 值无效: {weight}"
                })
                continue
        except (ValueError, TypeError):
            skipped.append({
                "row": str(row),
                "ticker": ticker,
                "reason": f"Weight 无法转换为数字: {weight}"
            })
            continue
        
        valid_holdings.append({
            "ticker": ticker.upper(),
            "weight": weight_float
        })
    
    return valid_holdings, skipped


def _normalize_symbol(value):
    if value is None:
        return None
    text = str(value).strip().upper()
    if not text:
        return None
    if not re.match(r'^[A-Z][A-Z0-9.\-]*$', text):
        return None
    return text


def _canonical_symbol_key(symbol: str) -> str:
    normalized = str(symbol or "").strip().upper()
    matched = _SHARE_CLASS_ALIAS_PATTERN.match(normalized)
    if not matched:
        return normalized
    return f"{matched.group(1)}.{matched.group(2)}"


def parse_etf_symbols(raw_etfs: str) -> list:
    symbols = []
    seen = set()
    for token in str(raw_etfs or "").split(','):
        symbol = token.strip().upper()
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        symbols.append(symbol)
    if not symbols:
        raise ValueError("ETF 列表为空，请使用 -s \"XLK,XLC,XLV\" 形式输入")
    return symbols


def parse_mc_coverage(raw_coverage: str) -> tuple:
    text = str(raw_coverage or "").strip().lower().replace(" ", "")
    if not text:
        raise ValueError("覆盖范围不能为空，请使用 -w t-10 或 -w 85")

    matched = re.match(r'^t-(\d+)$', text)
    if matched:
        value = int(matched.group(1))
        if value <= 0:
            raise ValueError("top 覆盖范围必须大于 0")
        return "top", value, f"top{value}"

    matched = re.match(r'^(\d+)$', text)
    if matched:
        value = int(matched.group(1))
        if value <= 0:
            raise ValueError("weight 覆盖范围必须大于 0")
        return "weight", value, f"weight{value}"

    matched = re.match(r'^top(\d+)$', text)
    if matched:
        value = int(matched.group(1))
        if value <= 0:
            raise ValueError("top 覆盖范围必须大于 0")
        return "top", value, f"top{value}"

    matched = re.match(r'^weight(\d+)$', text)
    if matched:
        value = int(matched.group(1))
        if value <= 0:
            raise ValueError("weight 覆盖范围必须大于 0")
        return "weight", value, f"weight{value}"

    raise ValueError(f"无效覆盖范围: {raw_coverage}。仅支持 -w t-10 或 -w 85")


def parse_refresh_holdings_coverage(raw_coverage: str) -> tuple:
    text = str(raw_coverage or "").strip().lower().replace(" ", "")
    if text == "all":
        return "all", 0, "all"
    return parse_mc_coverage(text)


def _normalize_api_base_url(raw_url: str) -> str:
    base_url = str(raw_url or DEFAULT_API_BASE_URL).strip()
    if not base_url:
        raise ValueError("API 地址不能为空")
    return base_url.rstrip("/")


def _http_json_request(method: str, url: str, payload=None, timeout: int = 10) -> dict:
    headers = {"Accept": "application/json"}
    body = None
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"

    req = urllib_request.Request(url=url, data=body, method=method.upper(), headers=headers)
    try:
        with urllib_request.urlopen(req, timeout=max(1, int(timeout))) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib_error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="ignore")
        detail = raw.strip()
        if detail:
            try:
                payload = json.loads(detail)
                if isinstance(payload, dict):
                    detail = payload.get("detail") or payload.get("message") or detail
            except Exception:
                pass
        raise RuntimeError(f"API 请求失败 ({exc.code}): {detail or exc.reason}") from exc
    except urllib_error.URLError as exc:
        raise RuntimeError(
            "无法连接后端 API，请先启动 FastAPI 服务，例如 `cd backend && .venv/bin/python -m uvicorn app.main:app --port 8000`"
        ) from exc


def _print_refresh_job_summary(job: dict, *, command_hint: str) -> None:
    job_id = job.get("id")
    queue_position = job.get("queue_position")
    message = job.get("message") or "任务已入队"

    print(message)
    print(f"Job ID: {job_id}")
    print(f"状态: {job.get('status')}")
    if queue_position is not None:
        print(f"队列位置: {queue_position}")
    print(f"查询状态: {command_hint}")


def _print_refresh_job_failure_details(job: dict) -> None:
    result = job.get("result")
    if not isinstance(result, dict):
        return
    items = result.get("items")
    if not isinstance(items, list) or not items:
        return

    failed_items = []
    for item in items:
        if not isinstance(item, dict):
            continue
        status = str(item.get("status") or "").strip().lower()
        if status in {"success", "snapshot"}:
            continue
        failed_items.append(item)

    if not failed_items:
        return

    print("失败明细:")
    for item in failed_items:
        symbol = str(item.get("symbol") or item.get("ticker") or "-").upper()
        coverage = str(item.get("coverage") or "").strip()
        label = f"{symbol} ({coverage})" if coverage else symbol
        message = item.get("message") or item.get("error") or item.get("status") or "未知错误"
        print(f"- {label}: {message}")


def _refresh_job_has_failures(job: dict) -> bool:
    if not isinstance(job, dict):
        return False
    if job.get("error"):
        return True
    try:
        if int(job.get("progress_failed") or 0) > 0:
            return True
    except (TypeError, ValueError):
        pass
    result = job.get("result")
    if isinstance(result, dict):
        summary_status = str(result.get("summary_status") or "").strip().lower()
        if summary_status in {"failed", "partial_success"}:
            return True
    return False


def _poll_refresh_job_until_terminal(
    job_id: int,
    *,
    api_base: str,
    timeout: int,
    max_wait_seconds: float = REFRESH_POST_SUBMIT_POLL_SECONDS,
    poll_interval_seconds: float = REFRESH_POST_SUBMIT_POLL_INTERVAL_SECONDS,
) -> dict | None:
    if not job_id:
        return None

    deadline = time.monotonic() + max(0.0, float(max_wait_seconds))
    while True:
        remaining = deadline - time.monotonic()
        if remaining <= 0:
            return None
        time.sleep(min(max(0.0, float(poll_interval_seconds)), remaining))
        try:
            job = _http_json_request(
                "GET",
                f"{api_base}/api/refresh-jobs/{job_id}",
                timeout=timeout,
            )
        except RuntimeError:
            return None

        status = str(job.get("status") or "").strip().lower()
        if _refresh_job_has_failures(job):
            return job
        if status in REFRESH_JOB_TERMINAL_STATUSES:
            return job


def cmd_refresh_etfs(args):
    symbols = parse_etf_symbols(args.symbols)
    api_base = _normalize_api_base_url(args.api_base)
    response = _http_json_request(
        "POST",
        f"{api_base}/api/refresh-jobs/etfs",
        payload={
            "symbols": symbols,
            "source": "cli",
        },
        timeout=args.timeout,
    )
    job = response.get("job") or {}
    _print_refresh_job_summary(
        job,
        command_hint=f"python -m app.cli refresh status {job.get('id')} --api-base {api_base}",
    )


def cmd_refresh_holdings(args):
    symbols = parse_etf_symbols(args.symbols)
    coverage_type, coverage_value, coverage_label = parse_refresh_holdings_coverage(args.coverage)
    api_base = _normalize_api_base_url(args.api_base)
    response = _http_json_request(
        "POST",
        f"{api_base}/api/refresh-jobs/holdings",
        payload={
            "items": [
                {
                    "symbol": symbol,
                    "coverage_type": coverage_type,
                    "coverage_value": coverage_value,
                    "related_etf_symbols": [],
                }
                for symbol in symbols
            ],
            "source": "cli",
        },
        timeout=args.timeout,
    )
    job = response.get("job") or {}
    command_hint = f"python -m app.cli refresh status {job.get('id')} --api-base {api_base}"
    print(f"已提交 holdings refresh: {','.join(symbols)} ({coverage_label})")
    _print_refresh_job_summary(
        job,
        command_hint=command_hint,
    )
    terminal_job = _poll_refresh_job_until_terminal(
        int(job.get("id") or 0),
        api_base=api_base,
        timeout=args.timeout,
    )
    if terminal_job and _refresh_job_has_failures(terminal_job):
        job_status = str(terminal_job.get("status") or "").strip().lower()
        if job_status in REFRESH_JOB_TERMINAL_STATUSES:
            print("后台任务快速返回错误:")
        else:
            print("后台任务已发现错误，任务仍在后台继续:")
        if terminal_job.get("message"):
            print(f"消息: {terminal_job.get('message')}")
        if terminal_job.get("error"):
            print(f"错误: {terminal_job.get('error')}")
        _print_refresh_job_failure_details(terminal_job)
        print(f"查询状态: {command_hint}")


def cmd_refresh_status(args):
    api_base = _normalize_api_base_url(args.api_base)
    job = _http_json_request(
        "GET",
        f"{api_base}/api/refresh-jobs/{args.job_id}",
        timeout=args.timeout,
    )
    print(f"Job ID: {job.get('id')}")
    print(f"类型: {job.get('job_type')}")
    print(f"状态: {job.get('status')}")
    print(f"进度: {job.get('progress_completed')}/{job.get('progress_total')} (失败 {job.get('progress_failed')})")
    if job.get("current_item"):
        print(f"当前标的: {job.get('current_item')}")
    if job.get("queue_position") is not None:
        print(f"队列位置: {job.get('queue_position')}")
    if job.get("message"):
        print(f"消息: {job.get('message')}")
    if job.get("error"):
        print(f"错误: {job.get('error')}")
    _print_refresh_job_failure_details(job)
    if args.show_result and job.get("result") is not None:
        print("结果:")
        print(json.dumps(job.get("result"), ensure_ascii=False, indent=2))


def cmd_refresh_list(args):
    api_base = _normalize_api_base_url(args.api_base)
    query_params = [f"limit={int(args.limit)}"]
    if args.status:
        query_params.append(f"status={args.status}")
    response = _http_json_request(
        "GET",
        f"{api_base}/api/refresh-jobs?{'&'.join(query_params)}",
        timeout=args.timeout,
    )
    items = response.get("items") or []
    if not items:
        print("暂无 refresh jobs")
        return

    print(f"{'ID':<6} {'TYPE':<10} {'STATUS':<12} {'PROGRESS':<14} {'QUEUE':<8} MESSAGE")
    for item in items:
        progress = f"{item.get('progress_completed', 0)}/{item.get('progress_total', 0)}"
        queue_position = item.get("queue_position")
        queue_text = "-" if queue_position is None else str(queue_position)
        message = str(item.get("message") or "")
        print(
            f"{str(item.get('id')):<6} "
            f"{str(item.get('job_type')):<10} "
            f"{str(item.get('status')):<12} "
            f"{progress:<14} "
            f"{queue_text:<8} "
            f"{message}"
        )


def load_mc_json_rows(file_path: str) -> list:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except FileNotFoundError:
        raise ValueError(f"文件不存在: {file_path}")
    except json.JSONDecodeError:
        raise ValueError(f"JSON 格式无效: {file_path}")

    if isinstance(payload, list):
        rows = payload
    elif isinstance(payload, dict) and isinstance(payload.get("data"), list):
        rows = payload.get("data")
    else:
        raise ValueError("JSON 根节点必须是数组，或包含 data 数组")

    dict_rows = [row for row in rows if isinstance(row, dict)]
    if not dict_rows:
        raise ValueError("JSON 中没有可用的数据行")
    return dict_rows


def load_finviz_source_rows(file_path: str) -> tuple[list, str]:
    file_ext = Path(file_path).suffix.lower()

    if file_ext == ".csv":
        try:
            import csv

            with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.DictReader(f)
                rows = [dict(row) for row in reader if isinstance(row, dict)]
        except FileNotFoundError:
            raise ValueError(f"文件不存在: {file_path}")
        except Exception as exc:
            raise ValueError(f"CSV 读取失败: {exc}")

        if not rows:
            raise ValueError("CSV 中没有可用的数据行")
        return rows, "csv"

    if file_ext == ".json":
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
        except FileNotFoundError:
            raise ValueError(f"文件不存在: {file_path}")
        except json.JSONDecodeError:
            raise ValueError(f"JSON 格式无效: {file_path}")

        if isinstance(payload, list):
            rows = payload
        elif isinstance(payload, dict) and isinstance(payload.get("data"), list):
            rows = payload.get("data")
        else:
            raise ValueError("JSON 根节点必须是数组，或包含 data 数组")

        dict_rows = [row for row in rows if isinstance(row, dict)]
        if not dict_rows:
            raise ValueError("JSON 中没有可用的数据行")
        return dict_rows, "json"

    raise ValueError("Finviz 文件仅支持 .json 或 .csv")


def filter_mc_source_fields(rows: list) -> list:
    filtered_rows = []
    for row in rows:
        symbol = _normalize_symbol(
            row.get("symbol")
            or row.get("Symbol")
            or row.get("Ticker")
            or row.get("ticker")
        )
        if not symbol:
            continue

        filtered = {}
        for field in MC_IMPORT_FIELDS:
            if field in row:
                filtered[field] = row.get(field)
        filtered["symbol"] = symbol
        filtered_rows.append(filtered)

    return filtered_rows


def _pick_coverage_symbols_for_cli(db, etf_symbol: str, coverage_type: str, coverage_value: int, import_date: date) -> tuple:
    from app.models.database import ETFHolding
    from sqlalchemy import func

    normalized_etf = str(etf_symbol or "").strip().upper()
    if not normalized_etf:
        return [], None

    latest_date = db.query(func.max(ETFHolding.data_date)).filter(
        ETFHolding.etf_symbol == normalized_etf,
        ETFHolding.data_date <= import_date
    ).scalar()

    if not latest_date:
        latest_date = db.query(func.max(ETFHolding.data_date)).filter(
            ETFHolding.etf_symbol == normalized_etf
        ).scalar()

    if not latest_date:
        return [], None

    holdings = db.query(ETFHolding).filter(
        ETFHolding.etf_symbol == normalized_etf,
        ETFHolding.data_date == latest_date
    ).order_by(ETFHolding.weight.desc()).all()

    if not holdings:
        return [], latest_date

    if coverage_type == "top":
        selected = holdings[:coverage_value]
    else:
        selected = []
        total_weight = 0.0
        for holding in holdings:
            selected.append(holding)
            total_weight += float(holding.weight or 0.0)
            if total_weight >= coverage_value:
                break

    symbols = []
    seen = set()
    for holding in selected:
        ticker = _normalize_symbol(getattr(holding, "ticker", None))
        if not ticker or ticker in seen:
            continue
        seen.add(ticker)
        symbols.append(ticker)

    return symbols, latest_date


def _select_rows_by_expected_symbols(rows: list, expected_symbols: list) -> tuple:
    expected_by_key = {}
    for symbol in expected_symbols:
        canonical = _canonical_symbol_key(symbol)
        if canonical and canonical not in expected_by_key:
            expected_by_key[canonical] = symbol

    matched_rows = {}
    duplicate_symbols = set()
    for row in rows:
        symbol = _normalize_symbol(row.get("symbol"))
        if not symbol:
            continue
        canonical = _canonical_symbol_key(symbol)
        resolved = expected_by_key.get(canonical)
        if not resolved:
            continue

        if resolved in matched_rows:
            duplicate_symbols.add(resolved)
            continue

        row_copy = dict(row)
        row_copy["symbol"] = resolved
        matched_rows[resolved] = row_copy

    selected_rows = [matched_rows[s] for s in expected_symbols if s in matched_rows]
    missing_symbols = [s for s in expected_symbols if s not in matched_rows]
    return selected_rows, missing_symbols, sorted(duplicate_symbols)


def _is_sqlite_locked_error(exc: Exception) -> bool:
    message = str(exc or "").lower()
    return "database is locked" in message or "database table is locked" in message


def _run_with_sqlite_lock_retry(action, action_label: str):
    last_exc = None
    for attempt in range(1, SQLITE_LOCK_RETRY_ATTEMPTS + 1):
        try:
            return action()
        except Exception as exc:
            last_exc = exc
            if not _is_sqlite_locked_error(exc) or attempt >= SQLITE_LOCK_RETRY_ATTEMPTS:
                raise
            wait_seconds = 0.5 * attempt
            print(
                f"提示: {action_label} 遇到数据库锁，"
                f"{wait_seconds:.1f} 秒后重试 ({attempt}/{SQLITE_LOCK_RETRY_ATTEMPTS})"
            )
            time.sleep(wait_seconds)
    raise last_exc


def _upsert_marketchameleon_rows(db, rows: list, import_date: date) -> tuple:
    from app.models.database import ImportedData

    inserted = 0
    updated = 0
    deduped_rows = {}

    for row in rows:
        symbol = _normalize_symbol(row.get("symbol"))
        if not symbol:
            continue
        canonical_symbol = _canonical_symbol_key(symbol)
        payload = dict(row)
        payload["symbol"] = canonical_symbol
        deduped_rows[canonical_symbol] = payload

    for symbol, payload in deduped_rows.items():
        existing = db.query(ImportedData).filter(
            ImportedData.symbol == symbol,
            ImportedData.date == import_date,
            ImportedData.source == "marketchameleon"
        ).first()

        if existing:
            existing.data = payload
            existing.date = import_date
            existing.created_at = datetime.now(timezone.utc).replace(tzinfo=None)
            updated += 1
        else:
            db.add(ImportedData(
                symbol=symbol,
                date=import_date,
                source="marketchameleon",
                data=payload
            ))
            inserted += 1

    return inserted, updated


def _upsert_finviz_rows(db, rows: list, import_date: date) -> tuple:
    from app.models.database import ImportedData

    inserted = 0
    updated = 0
    deduped_rows = {}
    duplicate_symbols = set()

    for row in rows:
        symbol = _normalize_symbol(row.get("symbol"))
        if not symbol:
            continue
        canonical_symbol = _canonical_symbol_key(symbol)
        payload = dict(row)
        payload["symbol"] = canonical_symbol
        if canonical_symbol in deduped_rows:
            duplicate_symbols.add(canonical_symbol)
        deduped_rows[canonical_symbol] = payload

    for symbol, payload in deduped_rows.items():
        existing = db.query(ImportedData).filter(
            ImportedData.symbol == symbol,
            ImportedData.date == import_date,
            ImportedData.source == "finviz"
        ).first()

        if existing:
            existing.data = payload
            existing.date = import_date
            existing.created_at = datetime.now(timezone.utc).replace(tzinfo=None)
            updated += 1
        else:
            db.add(ImportedData(
                symbol=symbol,
                date=import_date,
                source="finviz",
                data=payload
            ))
            inserted += 1

    return inserted, updated, sorted(duplicate_symbols), len(deduped_rows)


def cmd_import_finviz(args):
    """处理 finviz 命令"""
    from app.models.database import SessionLocal, ETF, init_db
    from app.services.parsers.finviz_parser import (
        calculate_breadth_metrics,
        get_summary_statistics,
        parse_finviz_json,
        validate_finviz_data,
    )

    init_db()

    data_date_str = args.date if args.date else _today_iso_date()
    file_path = args.file

    try:
        import_date = datetime.strptime(data_date_str, "%Y-%m-%d").date()
    except ValueError:
        print("错误: 日期格式无效，请使用 YYYY-MM-DD 格式")
        sys.exit(1)

    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        sys.exit(1)

    try:
        source_rows, source_format = load_finviz_source_rows(file_path)
    except ValueError as exc:
        print(f"错误: {exc}")
        sys.exit(1)

    coverage_mode = bool(args.coverage or args.etfs)
    try:
        if coverage_mode:
            if not args.coverage or not args.etfs:
                raise ValueError("Finviz 的 -s 与 -w 必须同时提供，例如 -s \"XLK,XLC\" -w 85")
            etf_symbols = parse_etf_symbols(args.etfs)
            coverage_type, coverage_value, coverage_label = parse_mc_coverage(args.coverage)
        else:
            etf_symbols = []
            coverage_type = None
            coverage_value = None
            coverage_label = None
    except ValueError as exc:
        print(f"错误: {exc}")
        sys.exit(1)

    parsed_rows = parse_finviz_json(source_rows)
    validation = validate_finviz_data(parsed_rows)
    breadth = calculate_breadth_metrics(parsed_rows)
    statistics = get_summary_statistics(parsed_rows)

    if not parsed_rows:
        print("错误: 源文件中没有可用的 Finviz 记录")
        sys.exit(1)

    print(f"\n{'='*70}")
    print("导入 Finviz 数据")
    print(f"{'='*70}")
    print(f"导入日期: {import_date}")
    print(f"文件: {file_path}")
    print(f"格式: {source_format.upper()}")
    if coverage_mode:
        print(f"ETF 列表: {', '.join(etf_symbols)}")
        print(f"覆盖范围: {coverage_label}")
    print(f"源数据行数: {len(source_rows)}")
    print(f"解析后行数: {len(parsed_rows)}")
    print(f"数据有效: {'是' if validation.get('is_valid') else '否'}")
    print(f"{'='*70}\n")

    warnings = validation.get("warnings") or []
    missing_fields = validation.get("missing_required_fields") or []
    if missing_fields:
        print(f"提示: 关键字段缺失 -> {', '.join(missing_fields)}")
    if warnings:
        for warning in warnings[:5]:
            print(f"提示: {warning}")
        if len(warnings) > 5:
            print(f"提示: 还有 {len(warnings) - 5} 条校验告警未展示")

    if not coverage_mode:
        def _write_all_rows():
            db = SessionLocal()
            try:
                inserted, updated, duplicate_symbols, unique_symbols = _upsert_finviz_rows(
                    db=db,
                    rows=parsed_rows,
                    import_date=import_date
                )
                db.commit()
                return inserted, updated, duplicate_symbols, unique_symbols
            except Exception:
                db.rollback()
                raise
            finally:
                db.close()

        try:
            inserted, updated, duplicate_symbols, unique_symbols = _run_with_sqlite_lock_retry(
                _write_all_rows,
                action_label="Finviz 导入"
            )
        except Exception as exc:
            print(f"\n错误: 导入失败 - {exc}")
            sys.exit(1)

        if duplicate_symbols:
            preview = ", ".join(duplicate_symbols[:10])
            suffix = "" if len(duplicate_symbols) <= 10 else f" ... 共 {len(duplicate_symbols)} 个"
            print(f"提示: 文件中存在重复标的，已按最后一条记录覆盖: {preview}{suffix}")

        print(f"\n{'='*70}")
        print("导入完成")
        print(f"{'='*70}")
        print(f"唯一标的: {unique_symbols}")
        print(f"新增记录: {inserted}")
        print(f"更新记录: {updated}")
        print(f"50MA 之上占比: {breadth.get('pct_above_sma50', 0):.1%}")
        print(f"200MA 之上占比: {breadth.get('pct_above_sma200', 0):.1%}")
        if statistics.get("avg_price") is not None:
            print(f"平均价格: {statistics['avg_price']:.2f}")
        print(f"{'='*70}\n")
        return

    read_db = SessionLocal()
    import_plans = []
    total_missing = 0
    etf_skipped = 0
    try:
        for etf_symbol in etf_symbols:
            expected_symbols, holdings_date = _pick_coverage_symbols_for_cli(
                db=read_db,
                etf_symbol=etf_symbol,
                coverage_type=coverage_type,
                coverage_value=coverage_value,
                import_date=import_date
            )

            if not expected_symbols:
                print(f"[{etf_symbol}] 跳过: 未找到可用持仓数据")
                etf_skipped += 1
                continue

            selected_rows, missing_symbols, duplicate_symbols = _select_rows_by_expected_symbols(
                parsed_rows,
                expected_symbols
            )

            if duplicate_symbols:
                preview = ", ".join(duplicate_symbols[:8])
                suffix = "" if len(duplicate_symbols) <= 8 else f" ... 共 {len(duplicate_symbols)} 个"
                print(f"[{etf_symbol}] 警告: 源文件存在重复标的，按首条记录导入: {preview}{suffix}")

            if missing_symbols:
                total_missing += len(missing_symbols)
                preview = ", ".join(missing_symbols[:10])
                suffix = "" if len(missing_symbols) <= 10 else f" ... 共 {len(missing_symbols)} 个"
                print(f"[{etf_symbol}] 提示: 覆盖范围缺失 {len(missing_symbols)} 个标的: {preview}{suffix}")

            if not selected_rows:
                print(f"[{etf_symbol}] 跳过: 未匹配到任何目标标的")
                etf_skipped += 1
                continue

            import_plans.append({
                "etf_symbol": etf_symbol,
                "holdings_date": holdings_date,
                "expected_count": len(expected_symbols),
                "selected_rows": selected_rows,
            })
    except Exception as exc:
        print(f"\n错误: 导入失败 - {exc}")
        sys.exit(1)
    finally:
        read_db.close()

    def _write_coverage_rows():
        db = SessionLocal()
        try:
            results = []
            for plan in import_plans:
                inserted, updated, _, _ = _upsert_finviz_rows(
                    db=db,
                    rows=plan["selected_rows"],
                    import_date=import_date
                )
                db.flush()

                etf_record = db.query(ETF).filter(ETF.symbol == plan["etf_symbol"]).first()
                if etf_record:
                    existing_ranges = getattr(etf_record, 'coverage_ranges', None) or []
                    if coverage_label not in existing_ranges:
                        existing_ranges.append(coverage_label)
                        etf_record.coverage_ranges = existing_ranges
                        etf_record.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)

                results.append({
                    "etf_symbol": plan["etf_symbol"],
                    "inserted": inserted,
                    "updated": updated,
                })
            db.commit()
            return results
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    try:
        write_results = _run_with_sqlite_lock_retry(
            _write_coverage_rows,
            action_label="Finviz 覆盖导入"
        )
    except Exception as exc:
        print(f"\n错误: 导入失败 - {exc}")
        sys.exit(1)

    result_by_etf = {item["etf_symbol"]: item for item in write_results}
    total_inserted = 0
    total_updated = 0
    total_matched = 0
    etf_success = 0
    unique_symbols = set()
    for plan in import_plans:
        result = result_by_etf.get(plan["etf_symbol"])
        if not result:
            continue
        total_inserted += result["inserted"]
        total_updated += result["updated"]
        total_matched += len(plan["selected_rows"])
        for item in plan["selected_rows"]:
            symbol = _normalize_symbol(item.get("symbol"))
            if symbol:
                unique_symbols.add(symbol)
        etf_success += 1

        print(
            f"[{plan['etf_symbol']}] 成功: 持仓日期={plan['holdings_date']}, 覆盖={plan['expected_count']}, "
            f"匹配={len(plan['selected_rows'])}, 新增={result['inserted']}, 更新={result['updated']}"
        )

    print(f"\n{'='*70}")
    print("导入完成")
    print(f"{'='*70}")
    print(f"成功 ETF: {etf_success}")
    print(f"跳过 ETF: {etf_skipped}")
    print(f"匹配记录: {total_matched}")
    print(f"唯一标的: {len(unique_symbols)}")
    print(f"缺失标的: {total_missing}")
    print(f"新增记录: {total_inserted}")
    print(f"更新记录: {total_updated}")
    print(f"{'='*70}\n")


def cmd_import_mc(args):
    """处理 mc 命令"""
    from app.models.database import SessionLocal, ETF, init_db
    from app.services.parsers.mc_parser import process_mc_data

    init_db()

    data_date_str = args.date if args.date else _today_iso_date()
    file_path = args.file

    try:
        import_date = datetime.strptime(data_date_str, "%Y-%m-%d").date()
    except ValueError:
        print("错误: 日期格式无效，请使用 YYYY-MM-DD 格式")
        sys.exit(1)

    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        sys.exit(1)

    try:
        source_rows = load_mc_json_rows(file_path)
        coverage_mode = bool(args.coverage or args.etfs)
        if coverage_mode:
            if not args.coverage or not args.etfs:
                raise ValueError("MarketChameleon 的 -s 与 -w 必须同时提供，例如 -s \"XLK,XLC\" -w 85")
            etf_symbols = parse_etf_symbols(args.etfs)
            coverage_type, coverage_value, coverage_label = parse_mc_coverage(args.coverage)
            filtered_source_rows = filter_mc_source_fields(source_rows)
        else:
            etf_symbols = []
            coverage_type = None
            coverage_value = None
            coverage_label = None
            filtered_source_rows = []
    except ValueError as exc:
        print(f"错误: {exc}")
        sys.exit(1)

    if coverage_mode and not filtered_source_rows:
        print("错误: 源文件中没有可用的 MarketChameleon 记录")
        sys.exit(1)

    processed_all_rows = process_mc_data(source_rows)
    if not processed_all_rows:
        print("错误: 源文件中没有可用的 MarketChameleon 记录")
        sys.exit(1)

    print(f"\n{'='*70}")
    print("导入 MarketChameleon 数据")
    print(f"{'='*70}")
    print(f"导入日期: {import_date}")
    print(f"文件: {file_path}")
    print(f"源数据行数: {len(source_rows)}")
    print(f"解析后行数: {len(processed_all_rows)}")
    if coverage_mode:
        print(f"ETF 列表: {', '.join(etf_symbols)}")
        print(f"覆盖范围: {coverage_label}")
        print(f"过滤后行数: {len(filtered_source_rows)}")
        print(f"保留字段: {', '.join(MC_IMPORT_FIELDS)}")
    print(f"{'='*70}\n")

    if not coverage_mode:
        def _write_all_rows():
            db = SessionLocal()
            try:
                inserted, updated = _upsert_marketchameleon_rows(
                    db=db,
                    rows=processed_all_rows,
                    import_date=import_date,
                )
                db.commit()
                return inserted, updated
            except Exception:
                db.rollback()
                raise
            finally:
                db.close()

        try:
            inserted, updated = _run_with_sqlite_lock_retry(
                _write_all_rows,
                action_label="MarketChameleon ETF 导入"
            )
        except Exception as exc:
            print(f"\n错误: 导入失败 - {exc}")
            sys.exit(1)

        heat_distribution = {}
        for item in processed_all_rows:
            heat_type = str(item.get("heat_type") or "normal").strip().lower() or "normal"
            heat_distribution[heat_type] = heat_distribution.get(heat_type, 0) + 1

        print(f"\n{'='*70}")
        print("导入完成")
        print(f"{'='*70}")
        print(f"唯一标的: {len({str(item.get('symbol', '')).upper() for item in processed_all_rows if item.get('symbol')})}")
        print(f"新增记录: {inserted}")
        print(f"更新记录: {updated}")
        if heat_distribution:
            ordered_heat = ", ".join(f"{key}={value}" for key, value in sorted(heat_distribution.items()))
            print(f"热度分布: {ordered_heat}")
        print(f"{'='*70}\n")
        return

    read_db = SessionLocal()
    import_plans = []
    total_missing = 0
    etf_skipped = 0
    try:
        for etf_symbol in etf_symbols:
            expected_symbols, holdings_date = _pick_coverage_symbols_for_cli(
                db=read_db,
                etf_symbol=etf_symbol,
                coverage_type=coverage_type,
                coverage_value=coverage_value,
                import_date=import_date
            )

            if not expected_symbols:
                print(f"[{etf_symbol}] 跳过: 未找到可用持仓数据")
                etf_skipped += 1
                continue

            selected_rows, missing_symbols, duplicate_symbols = _select_rows_by_expected_symbols(
                filtered_source_rows,
                expected_symbols
            )

            if duplicate_symbols:
                preview = ", ".join(duplicate_symbols[:8])
                suffix = "" if len(duplicate_symbols) <= 8 else f" ... 共 {len(duplicate_symbols)} 个"
                print(f"[{etf_symbol}] 警告: 源文件存在重复标的，按首条记录导入: {preview}{suffix}")

            if missing_symbols:
                total_missing += len(missing_symbols)
                preview = ", ".join(missing_symbols[:10])
                suffix = "" if len(missing_symbols) <= 10 else f" ... 共 {len(missing_symbols)} 个"
                print(f"[{etf_symbol}] 提示: 覆盖范围缺失 {len(missing_symbols)} 个标的: {preview}{suffix}")

            if not selected_rows:
                print(f"[{etf_symbol}] 跳过: 未匹配到任何目标标的")
                etf_skipped += 1
                continue

            processed_rows = process_mc_data(selected_rows)
            import_plans.append({
                "etf_symbol": etf_symbol,
                "holdings_date": holdings_date,
                "expected_count": len(expected_symbols),
                "processed_rows": processed_rows,
            })
    except Exception as e:
        print(f"\n错误: 导入失败 - {e}")
        sys.exit(1)
    finally:
        read_db.close()

    def _write_mc_rows():
        db = SessionLocal()
        try:
            results = []
            for plan in import_plans:
                inserted, updated = _upsert_marketchameleon_rows(
                    db,
                    plan["processed_rows"],
                    import_date
                )
                db.flush()

                etf_record = db.query(ETF).filter(ETF.symbol == plan["etf_symbol"]).first()
                if etf_record:
                    existing_ranges = getattr(etf_record, 'coverage_ranges', None) or []
                    if coverage_label not in existing_ranges:
                        existing_ranges.append(coverage_label)
                        etf_record.coverage_ranges = existing_ranges
                        etf_record.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)

                results.append({
                    "etf_symbol": plan["etf_symbol"],
                    "inserted": inserted,
                    "updated": updated,
                })
            db.commit()
            return results
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

    try:
        write_results = _run_with_sqlite_lock_retry(
            _write_mc_rows,
            action_label="MarketChameleon 导入"
        )
    except Exception as e:
        print(f"\n错误: 导入失败 - {e}")
        sys.exit(1)

    result_by_etf = {item["etf_symbol"]: item for item in write_results}
    total_inserted = 0
    total_updated = 0
    total_matched = 0
    etf_success = 0
    unique_symbols = set()
    for plan in import_plans:
        result = result_by_etf.get(plan["etf_symbol"])
        if not result:
            continue
        total_inserted += result["inserted"]
        total_updated += result["updated"]
        total_matched += len(plan["processed_rows"])
        for item in plan["processed_rows"]:
            symbol = _normalize_symbol(item.get("symbol"))
            if symbol:
                unique_symbols.add(symbol)
        etf_success += 1

        print(
            f"[{plan['etf_symbol']}] 成功: 持仓日期={plan['holdings_date']}, 覆盖={plan['expected_count']}, "
            f"匹配={len(plan['processed_rows'])}, 新增={result['inserted']}, 更新={result['updated']}"
        )

    print(f"\n{'='*70}")
    print("导入完成")
    print(f"{'='*70}")
    print(f"成功 ETF: {etf_success}")
    print(f"跳过 ETF: {etf_skipped}")
    print(f"匹配记录: {total_matched}")
    print(f"唯一标的: {len(unique_symbols)}")
    print(f"缺失标的: {total_missing}")
    print(f"新增记录: {total_inserted}")
    print(f"更新记录: {total_updated}")
    print(f"{'='*70}\n")


def cmd_uploads(args):
    """处理 uploads 命令"""
    from app.models.database import (
        SessionLocal, ETF, ETFHolding, HoldingsUploadLog,
        is_valid_sector_symbol, VALID_SECTOR_SYMBOLS, init_db
    )
    
    # 初始化数据库
    init_db()
    
    # 验证参数
    etf_type = args.type
    etf_symbol = args.etf_symbol.upper()
    data_date_str = args.date if args.date else _today_iso_date()
    file_path = args.file
    parent_sector = args.sector.upper() if args.sector else None
    
    # 验证日期格式
    try:
        data_date = datetime.strptime(data_date_str, "%Y-%m-%d").date()
    except ValueError:
        print(f"错误: 日期格式无效，请使用 YYYY-MM-DD 格式")
        sys.exit(1)
    
    # 验证 ETF 类型
    if etf_type not in ["sector", "industry"]:
        print(f"错误: ETF 类型必须是 'sector' 或 'industry'")
        sys.exit(1)
    
    # 板块 ETF 验证
    if etf_type == "sector":
        if not is_valid_sector_symbol(etf_symbol):
            print(f"错误: 无效的板块 ETF 符号")
            print(f"有效的板块 ETF: {', '.join(VALID_SECTOR_SYMBOLS)}")
            sys.exit(1)
    
    # 行业 ETF 验证
    if etf_type == "industry" and parent_sector:
        if not is_valid_sector_symbol(parent_sector):
            print(f"错误: 无效的父板块符号")
            print(f"有效的板块 ETF: {', '.join(VALID_SECTOR_SYMBOLS)}")
            sys.exit(1)
    
    # 验证文件存在
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        sys.exit(1)
    
    # 验证文件类型
    if not file_path.endswith(('.xlsx', '.xls', '.csv')):
        print(f"错误: 只支持 xlsx、xls 或 csv 文件格式")
        sys.exit(1)
    
    print(f"\n{'='*60}")
    print(f"上传 ETF Holdings")
    print(f"{'='*60}")
    print(f"ETF 类型: {etf_type}")
    print(f"ETF 符号: {etf_symbol}")
    if parent_sector:
        print(f"父板块: {parent_sector}")
    print(f"数据日期: {data_date_str}")
    print(f"文件: {file_path}")
    print(f"{'='*60}\n")
    
    # 解析文件
    if file_path.endswith('.csv'):
        print("正在解析 csv 文件...")
        raw_holdings = parse_csv_holdings(file_path)
    else:
        print("正在解析 xlsx 文件...")
        raw_holdings = parse_xlsx_holdings(file_path)
    print(f"找到 {len(raw_holdings)} 行数据")
    
    # 验证数据
    print("正在验证数据...")
    valid_holdings, skipped = validate_holdings(raw_holdings)
    print(f"有效记录: {len(valid_holdings)} 条")
    print(f"跳过记录: {len(skipped)} 条")
    
    if skipped and len(skipped) <= 10:
        print("\n跳过的记录详情:")
        for s in skipped:
            print(f"  行 {s['row']}: {s['ticker']} - {s['reason']}")
    elif skipped:
        print(f"\n前 10 条跳过记录详情:")
        for s in skipped[:10]:
            print(f"  行 {s['row']}: {s['ticker']} - {s['reason']}")
        print(f"  ... 还有 {len(skipped) - 10} 条")
    
    if not valid_holdings:
        print("\n错误: 没有有效的持仓数据")
        sys.exit(1)
    
    # 写入数据库
    print("\n正在写入数据库...")
    db = SessionLocal()
    
    try:
        # 查找或创建 ETF
        etf = db.query(ETF).filter(ETF.symbol == etf_symbol).first()
        
        if not etf:
            etf = ETF(
                symbol=etf_symbol,
                name=etf_symbol,
                type=etf_type,
                parent_sector=parent_sector if etf_type == "industry" else None,
                score=0.0,
                rank=0,
                delta={"delta3d": None, "delta5d": None},
                completeness=0.0,
                holdings_count=0
            )
            db.add(etf)
            db.flush()
            print(f"创建新的 ETF 记录: {etf_symbol}")
        
        # 删除该 ETF 在指定日期的旧持仓数据
        deleted = db.query(ETFHolding).filter(
            ETFHolding.etf_id == etf.id,
            ETFHolding.data_date == data_date
        ).delete()
        if deleted:
            print(f"删除旧数据: {deleted} 条记录")
        
        # 插入新的持仓数据
        for h in valid_holdings:
            holding = ETFHolding(
                etf_id=etf.id,
                etf_symbol=etf_symbol,
                ticker=h["ticker"],
                weight=h["weight"],
                data_date=data_date
            )
            db.add(holding)
        
        # 更新 ETF 的持仓数量
        etf.holdings_count = len(valid_holdings)
        etf.updated_at = utc_now_naive()
        
        # 删除该 ETF 在指定日期的旧上传日志（支持重复上传）
        db.query(HoldingsUploadLog).filter(
            HoldingsUploadLog.etf_symbol == etf_symbol,
            HoldingsUploadLog.data_date == data_date
        ).delete()
        
        # 记录上传日志
        upload_log = HoldingsUploadLog(
            etf_symbol=etf_symbol,
            etf_type=etf_type,
            data_date=data_date,
            file_name=os.path.basename(file_path),
            records_count=len(valid_holdings),
            skipped_count=len(skipped),
            status="success"
        )
        db.add(upload_log)
        
        db.commit()
        
        print(f"\n{'='*60}")
        print(f"上传成功!")
        print(f"{'='*60}")
        print(f"ETF: {etf_symbol}")
        print(f"日期: {data_date_str}")
        print(f"导入记录: {len(valid_holdings)} 条")
        print(f"跳过记录: {len(skipped)} 条")
        print(f"{'='*60}\n")
        
    except Exception as e:
        db.rollback()
        print(f"\n错误: 写入数据库失败 - {e}")
        
        # 记录失败日志
        try:
            upload_log = HoldingsUploadLog(
                etf_symbol=etf_symbol,
                etf_type=etf_type,
                data_date=data_date,
                file_name=os.path.basename(file_path),
                records_count=0,
                skipped_count=0,
                status="error",
                error_message=str(e)
            )
            db.add(upload_log)
            db.commit()
        except:
            db.rollback()
        
        sys.exit(1)
    finally:
        db.close()


def cmd_init(args):
    """处理 init 命令"""
    from app.models.database import init_db, init_default_sector_etfs
    
    print("正在初始化数据库...")
    init_db()
    print("数据库表已创建")
    
    print("正在初始化默认板块 ETF...")
    init_default_sector_etfs()
    print("默认板块 ETF 已初始化")
    
    print("\n初始化完成!")


def cmd_list_etfs(args):
    """列出所有 ETF"""
    from app.models.database import SessionLocal, ETF, init_db
    
    init_db()
    db = SessionLocal()
    
    try:
        etf_type = args.type if args.type else None
        
        query = db.query(ETF)
        if etf_type:
            query = query.filter(ETF.type == etf_type)
        
        etfs = query.order_by(ETF.type, ETF.symbol).all()
        
        print(f"\n{'='*70}")
        print(f"ETF 列表 (共 {len(etfs)} 个)")
        print(f"{'='*70}")
        print(f"{'类型':<10} {'符号':<10} {'名称':<30} {'持仓数':<10}")
        print(f"{'-'*70}")
        
        for etf in etfs:
            print(f"{etf.type:<10} {etf.symbol:<10} {etf.name[:28]:<30} {etf.holdings_count:<10}")
        
        print(f"{'='*70}\n")
        
    finally:
        db.close()


def cmd_list_holdings(args):
    """列出 ETF 持仓"""
    from app.models.database import SessionLocal, ETF, ETFHolding, init_db
    from sqlalchemy import func
    
    init_db()
    db = SessionLocal()
    
    try:
        etf_symbol = args.etf_symbol.upper()
        
        etf = db.query(ETF).filter(ETF.symbol == etf_symbol).first()
        if not etf:
            print(f"错误: 未找到 ETF '{etf_symbol}'")
            sys.exit(1)
        
        # 获取最新日期
        if args.date:
            try:
                data_date = datetime.strptime(args.date, "%Y-%m-%d").date()
            except ValueError:
                print(f"错误: 日期格式无效，请使用 YYYY-MM-DD 格式")
                sys.exit(1)
        else:
            data_date = db.query(func.max(ETFHolding.data_date)).filter(
                ETFHolding.etf_id == etf.id
            ).scalar()
        
        if not data_date:
            print(f"错误: ETF '{etf_symbol}' 没有持仓数据")
            sys.exit(1)
        
        holdings = db.query(ETFHolding).filter(
            ETFHolding.etf_id == etf.id,
            ETFHolding.data_date == data_date
        ).order_by(ETFHolding.weight.desc()).all()
        
        print(f"\n{'='*50}")
        print(f"{etf_symbol} 持仓 ({data_date})")
        print(f"{'='*50}")
        print(f"{'#':<5} {'Ticker':<10} {'Weight (%)':<15}")
        print(f"{'-'*50}")
        
        for idx, h in enumerate(holdings, 1):
            print(f"{idx:<5} {h.ticker:<10} {h.weight:<15.2f}")
        
        print(f"{'-'*50}")
        print(f"总计: {len(holdings)} 个持仓")
        print(f"{'='*50}\n")
        
    finally:
        db.close()


def build_parser():
    parser = argparse.ArgumentParser(
        prog="python -m app.cli",
        description='Momentum Radar 命令行工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 上传板块 ETF holdings（日期可选，默认为当天）
  python -m app.cli uploads -t sector -a XLK holdings.xlsx
  python -m app.cli uploads -d 2026-01-25 -t sector -a XLK holdings.xlsx
  
  # 上传行业 ETF holdings（需要指定父板块）
  python -m app.cli uploads -d 2026-01-25 -t industry -s XLK -a SOXX holdings.xlsx
  
  # 更新数据（日期可选，默认为当天）
  python -m app.cli update -t sector -a XLE xle.xlsx
  python -m app.cli update -d 2026-01-28 -t sector -a XLE xle.xlsx
  
  # 初始化数据库
  python -m app.cli init
  
  # 列出所有 ETF
  python -m app.cli list-etfs
  
  # 列出 ETF 持仓
  python -m app.cli list-holdings XLK

  # 导入 Finviz ETF 数据（支持 JSON/CSV）
  python -m app.cli finviz etfs -f finviz_export.csv
  python -m app.cli finviz etfs -d 2026-03-06 -f finviz.json
  python -m app.cli finviz etfs -s "XLK,XLC,XLV" -w 85 -f finviz_export.csv

  # 导入 MarketChameleon ETF 数据（整文件）
  python -m app.cli mc etfs -f marketchameleon_etfs.json

  # 导入 MarketChameleon ETF 数据（按多个 ETF + 单一覆盖范围）
  python -m app.cli mc etfs -d 2026-03-06 -s "XLK,XLC,XLV" -w 85 -f marketchameleon.json
  python -m app.cli mc etfs -s "XLK,XLC,XLV" -w t-10 -f marketchameleon.json

  # 激活 backend 虚拟环境并执行 ./bin/install-cli-shortcuts 后，可直接使用短命令
  refresh etfs -s "XLK,XLF,SOXX"
  finviz -f export.csv
  mc -f marketchameleon.json

  # 后台提交 refresh 任务（命令不会等待刷新完成）
  python -m app.cli refresh etfs -s "XLK,XLF,SOXX"
  python -m app.cli refresh holdings -s "XLK,SOXX" -w t-20
  python -m app.cli refresh status 12

  # 旧写法仍兼容，会自动按 etfs 处理
  python -m app.cli finviz -f finviz_export.csv
  python -m app.cli mc -f marketchameleon_etfs.json
  python -m app.cli mc -s "XLK,XLC,XLV" -w 85 -f marketchameleon.json
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='可用命令')
    
    # uploads 命令
    uploads_parser = subparsers.add_parser('uploads', help='上传 ETF Holdings 文件 (xlsx/xls/csv)')
    uploads_parser.add_argument(
        '-d',
        '--date',
        required=False,
        default=_today_iso_date(),
        help='数据日期 (YYYY-MM-DD)，默认为当天',
    )
    uploads_parser.add_argument('-t', '--type', required=True, choices=['sector', 'industry'],
                               help='ETF 类型: sector 或 industry')
    uploads_parser.add_argument('-a', '--etf-symbol', required=True, dest='etf_symbol',
                               help='ETF 符号 (如 XLK, SOXX)')
    uploads_parser.add_argument('-s', '--sector', help='父板块符号 (仅 industry 类型需要)')
    uploads_parser.add_argument(
        '-f', '--file',
        required=False,
        help='Holdings 文件路径或文件名；仅填文件名时会按 cfg.yaml 的 cli.downloads_dir / holdings_file_prefix 自动补全'
    )
    uploads_parser.add_argument(
        'file_arg',
        nargs='?',
        help='兼容旧写法的 Holdings 文件路径 (xlsx/xls/csv)'
    )
    uploads_parser.set_defaults(func=cmd_uploads)
    
    # update 命令 (uploads 的别名，更语义化)
    update_parser = subparsers.add_parser('update', help='更新 ETF Holdings 数据 (与 uploads 相同)')
    update_parser.add_argument(
        '-d',
        '--date',
        required=False,
        default=_today_iso_date(),
        help='数据日期 (YYYY-MM-DD)，默认为当天',
    )
    update_parser.add_argument('-t', '--type', required=True, choices=['sector', 'industry'],
                              help='ETF 类型: sector 或 industry')
    update_parser.add_argument('-a', '--etf-symbol', required=True, dest='etf_symbol',
                              help='ETF 符号 (如 XLK, SOXX)')
    update_parser.add_argument('-s', '--sector', help='父板块符号 (仅 industry 类型需要)')
    update_parser.add_argument(
        '-f', '--file',
        required=False,
        help='Holdings 文件路径或文件名；仅填文件名时会按 cfg.yaml 的 cli.downloads_dir / holdings_file_prefix 自动补全'
    )
    update_parser.add_argument(
        'file_arg',
        nargs='?',
        help='兼容旧写法的 Holdings 文件路径 (xlsx/xls/csv)'
    )
    update_parser.set_defaults(func=cmd_uploads)
    
    # init 命令
    init_parser = subparsers.add_parser('init', help='初始化数据库和默认数据')
    init_parser.set_defaults(func=cmd_init)
    
    # list-etfs 命令
    list_etfs_parser = subparsers.add_parser('list-etfs', help='列出所有 ETF')
    list_etfs_parser.add_argument('-t', '--type', choices=['sector', 'industry'],
                                  help='筛选 ETF 类型')
    list_etfs_parser.set_defaults(func=cmd_list_etfs)
    
    # list-holdings 命令
    list_holdings_parser = subparsers.add_parser('list-holdings', help='列出 ETF 持仓')
    list_holdings_parser.add_argument('etf_symbol', help='ETF 符号')
    list_holdings_parser.add_argument('-d', '--date', help='数据日期 (默认最新)')
    list_holdings_parser.set_defaults(func=cmd_list_holdings)

    # finviz 命令
    finviz_parser = subparsers.add_parser(
        'finviz',
        help='导入 Finviz 数据'
    )
    finviz_subparsers = finviz_parser.add_subparsers(dest='resource')
    finviz_subparsers.required = True
    import_finviz_parser = finviz_subparsers.add_parser(
        PROVIDER_ETF_SUBCOMMAND,
        help='导入 Finviz ETF JSON/CSV 数据'
    )
    import_finviz_parser.add_argument(
        '-d', '--date', required=False,
        help='导入日期 (YYYY-MM-DD)，默认为当天'
    )
    import_finviz_parser.add_argument(
        '-f', '--file', required=True,
        help='Finviz JSON/CSV 文件路径或文件名；仅填文件名时会按 cfg.yaml 的 cli.downloads_dir / finviz_file_prefix 自动补全'
    )
    import_finviz_parser.add_argument(
        '-s', '--etfs', required=False,
        help='ETF 列表，逗号分隔（如 "XLK,XLC,XLV"）；与 -w 一起使用时按持仓覆盖范围筛选'
    )
    import_finviz_parser.add_argument(
        '-w', '--coverage', required=False,
        help='可选覆盖范围；与 mc 一致，支持 t-10(=top10) 或 85(=weight85)'
    )
    import_finviz_parser.set_defaults(resource=PROVIDER_ETF_SUBCOMMAND)
    import_finviz_parser.set_defaults(func=cmd_import_finviz)

    # mc 命令
    mc_parser = subparsers.add_parser(
        'mc',
        help='导入 MarketChameleon 数据'
    )
    mc_subparsers = mc_parser.add_subparsers(dest='resource')
    mc_subparsers.required = True
    import_mc_parser = mc_subparsers.add_parser(
        PROVIDER_ETF_SUBCOMMAND,
        help='按 ETF 覆盖范围导入 MarketChameleon ETF JSON 数据'
    )
    import_mc_parser.add_argument(
        '-d', '--date', required=False,
        help='导入日期 (YYYY-MM-DD)，默认为当天'
    )
    import_mc_parser.add_argument(
        '-f', '--file', required=True,
        help='MarketChameleon JSON 文件路径或文件名；仅填文件名时会按 cfg.yaml 的 cli.downloads_dir / mc_file_prefix 自动补全'
    )
    import_mc_parser.add_argument(
        '-s', '--etfs', required=False,
        help='可选 ETF 列表，逗号分隔（如 "XLK,XLC,XLV"）；与 -w 一起提供时按持仓覆盖范围筛选'
    )
    import_mc_parser.add_argument(
        '-w', '--coverage', required=False,
        help='可选覆盖范围；与 -s 一起提供时仅导入对应覆盖标的，支持 t-10(=top10) 或 85(=weight85)'
    )
    import_mc_parser.set_defaults(resource=PROVIDER_ETF_SUBCOMMAND)
    import_mc_parser.set_defaults(func=cmd_import_mc)

    refresh_parser = subparsers.add_parser(
        'refresh',
        help='提交后台 refresh 任务并由服务端串行执行'
    )
    refresh_subparsers = refresh_parser.add_subparsers(dest='resource')
    refresh_subparsers.required = True

    refresh_etfs_parser = refresh_subparsers.add_parser(
        'etfs',
        help='后台刷新多个 ETF，命令提交后立即返回'
    )
    refresh_etfs_parser.add_argument(
        '-s', '--symbols', required=True,
        help='ETF 列表，逗号分隔（如 "XLK,XLF,SOXX"）'
    )
    refresh_etfs_parser.add_argument(
        '--api-base', default=DEFAULT_API_BASE_URL,
        help=f'后端 API 地址，默认 {DEFAULT_API_BASE_URL}'
    )
    refresh_etfs_parser.add_argument(
        '--timeout', type=int, default=10,
        help='提交任务请求超时（秒）'
    )
    refresh_etfs_parser.set_defaults(func=cmd_refresh_etfs)

    refresh_holdings_parser = refresh_subparsers.add_parser(
        'holdings',
        help='后台串行刷新多个 ETF holdings，命令提交后立即返回'
    )
    refresh_holdings_parser.add_argument(
        '-s', '--symbols', required=True,
        help='ETF 列表，逗号分隔（如 "XLK,SOXX"）'
    )
    refresh_holdings_parser.add_argument(
        '-w', '--coverage', default='t-20',
        help='覆盖范围，支持 t-20 / 85 / all，默认 t-20'
    )
    refresh_holdings_parser.add_argument(
        '--api-base', default=DEFAULT_API_BASE_URL,
        help=f'后端 API 地址，默认 {DEFAULT_API_BASE_URL}'
    )
    refresh_holdings_parser.add_argument(
        '--timeout', type=int, default=10,
        help='提交任务请求超时（秒）'
    )
    refresh_holdings_parser.set_defaults(func=cmd_refresh_holdings)

    refresh_status_parser = refresh_subparsers.add_parser(
        'status',
        help='查看后台 refresh job 状态'
    )
    refresh_status_parser.add_argument('job_id', type=int, help='job id')
    refresh_status_parser.add_argument(
        '--api-base', default=DEFAULT_API_BASE_URL,
        help=f'后端 API 地址，默认 {DEFAULT_API_BASE_URL}'
    )
    refresh_status_parser.add_argument(
        '--timeout', type=int, default=10,
        help='请求超时（秒）'
    )
    refresh_status_parser.add_argument(
        '--show-result', action='store_true',
        help='输出完整 result JSON'
    )
    refresh_status_parser.set_defaults(func=cmd_refresh_status)

    refresh_list_parser = refresh_subparsers.add_parser(
        'list',
        help='列出最近的后台 refresh jobs'
    )
    refresh_list_parser.add_argument(
        '--status', choices=['pending', 'running', 'completed', 'failed'],
        help='按状态过滤'
    )
    refresh_list_parser.add_argument(
        '--limit', type=int, default=10,
        help='返回 job 数量，默认 10'
    )
    refresh_list_parser.add_argument(
        '--api-base', default=DEFAULT_API_BASE_URL,
        help=f'后端 API 地址，默认 {DEFAULT_API_BASE_URL}'
    )
    refresh_list_parser.add_argument(
        '--timeout', type=int, default=10,
        help='请求超时（秒）'
    )
    refresh_list_parser.set_defaults(func=cmd_refresh_list)

    return parser


def main():
    args = parse_cli_args()
    parser = build_parser()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)

    if args.command in {"uploads", "update", "init", "list-etfs", "list-holdings", "finviz", "mc"}:
        _ensure_db_dependencies()
    
    args.func(args)


if __name__ == '__main__':
    main()
