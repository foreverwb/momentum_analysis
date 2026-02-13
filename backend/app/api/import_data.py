"""
数据导入 API
Data Import API Endpoints

支持:
- Finviz 技术指标数据导入
- MarketChameleon 期权数据导入
- CSV/JSON 批量导入
- ETF Holdings xlsx 文件上传
"""

from fastapi import APIRouter, HTTPException, File, UploadFile, Form, Depends, Query
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Dict, Optional, Any, Tuple
from datetime import datetime, date
import json
import csv
import io
import logging
import re

from app.models import (
    get_db, ETF, ETFHolding, HoldingsUploadLog, ImportedData,
    is_valid_ticker, is_valid_sector_symbol, VALID_SECTOR_SYMBOLS
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/import", tags=["Import"])
_COVERAGE_PATTERN = re.compile(r"^(top|weight)(\d+)$", re.IGNORECASE)
_SHARE_CLASS_ALIAS_PATTERN = re.compile(r"^([A-Z][A-Z0-9]{0,5})[.-]([A-Z])$")


def _upsert_imported_data(db: Session, source: str, items: List[Dict[str, Any]]) -> int:
    """将导入数据写入 imported_data 表，按 symbol+date+source 去重。"""
    if not items:
        return 0
    today = date.today()
    count = 0
    for item in items:
        symbol = item.get("symbol") if isinstance(item, dict) else None
        if not symbol:
            continue
        normalized_symbol = _normalize_symbol(symbol)
        if not normalized_symbol:
            continue
        symbol = _canonical_symbol_key(normalized_symbol)
        payload = dict(item) if isinstance(item, dict) else {}
        payload["symbol"] = symbol
        existing = db.query(ImportedData).filter(
            ImportedData.symbol == symbol,
            ImportedData.date == today,
            ImportedData.source == source
        ).first()
        if existing:
            existing.data = payload
            existing.date = today
            # 复用记录时同步刷新导入时间，便于前端按“北京时间 08:00 起算”判定最新状态
            existing.created_at = datetime.utcnow()
        else:
            db.add(ImportedData(
                symbol=symbol,
                date=today,
                source=source,
                data=payload
            ))
        count += 1
    return count


# ==================== Pydantic Models ====================

class FinvizImportRequest(BaseModel):
    """Finviz 数据导入请求"""
    etf_symbol: str = Field(..., description="关联的 ETF 代码")
    coverage: str = Field("top20", description="覆盖范围: top10/top15/top20/top25/top30/weight60..weight85/all")
    data: List[Dict[str, Any]] = Field(..., description="Finviz 数据列表")


class FinvizImportResponse(BaseModel):
    """Finviz 导入响应"""
    status: str
    etf_symbol: str
    coverage: str
    records_imported: int
    breadth_metrics: Dict[str, Any]
    validation: Dict[str, Any]
    statistics: Optional[Dict[str, Any]] = None
    symbol_alias_mappings: Optional[List[str]] = None


class MCImportRequest(BaseModel):
    """MarketChameleon 数据导入请求"""
    etf_symbol: Optional[str] = Field(None, description="关联 ETF 代码（可选，提供时会进行覆盖范围校验）")
    coverage: Optional[str] = Field(None, description="覆盖范围（可选，提供时会进行覆盖范围校验，支持 all）")
    symbols: List[str] = Field(default=[], description="股票代码列表（可选，自动从数据中提取）")
    data: List[Dict[str, Any]] = Field(..., description="MarketChameleon 数据列表")


class MCImportResponse(BaseModel):
    """MarketChameleon 导入响应"""
    status: str
    records_imported: int
    heat_distribution: Dict[str, int]
    data: Optional[List[Dict[str, Any]]] = None


class BulkImportResponse(BaseModel):
    """批量导入响应"""
    status: str
    source: str
    records_imported: int
    errors: List[str]


class HoldingsImportRequest(BaseModel):
    """ETF 持仓导入请求"""
    etf_symbol: str
    holdings: List[Dict[str, Any]]


class HoldingsImportResponse(BaseModel):
    """ETF 持仓导入响应"""
    status: str
    etf_symbol: str
    holdings_count: int
    scored_holdings: Optional[List[Dict[str, Any]]] = None


class HoldingsUploadResponse(BaseModel):
    """Holdings xlsx 上传响应"""
    status: str
    etf_symbol: str
    etf_type: str
    data_date: str
    records_imported: int
    records_skipped: int
    skipped_details: Optional[List[Dict[str, str]]] = None


# ==================== 辅助函数 ====================

def _normalize_symbol(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    if not text:
        return None
    return text if is_valid_ticker(text) else None


def _canonical_symbol_key(symbol: str) -> str:
    normalized = str(symbol or "").strip().upper()
    matched = _SHARE_CLASS_ALIAS_PATTERN.match(normalized)
    if not matched:
        return normalized
    return f"{matched.group(1)}.{matched.group(2)}"


def _parse_coverage_selection(coverage: str) -> Tuple[str, int]:
    text = str(coverage or "").strip().lower()
    if text == "all":
        return "all", 0
    matched = _COVERAGE_PATTERN.match(text)
    if not matched:
        raise HTTPException(
            status_code=400,
            detail=f"覆盖范围格式无效: {coverage}。支持示例: top10, top20, weight70, weight85, all"
        )
    coverage_type = matched.group(1).lower()
    coverage_value = int(matched.group(2))
    if coverage_value <= 0:
        raise HTTPException(status_code=400, detail=f"覆盖范围值必须大于 0: {coverage}")
    return coverage_type, coverage_value


def _pick_coverage_symbols(
    db: Session,
    etf_symbol: str,
    coverage: str
) -> List[str]:
    normalized_etf = str(etf_symbol or "").strip().upper()
    if not normalized_etf:
        raise HTTPException(status_code=400, detail="etf_symbol 不能为空")

    coverage_type, coverage_value = _parse_coverage_selection(coverage)

    latest_date = db.query(func.max(ETFHolding.data_date)).filter(
        ETFHolding.etf_symbol == normalized_etf
    ).scalar()
    if not latest_date:
        raise HTTPException(
            status_code=400,
            detail=f"ETF {normalized_etf} 没有可用持仓数据，无法校验覆盖范围 {coverage}"
        )

    holdings = db.query(ETFHolding).filter(
        ETFHolding.etf_symbol == normalized_etf,
        ETFHolding.data_date == latest_date
    ).order_by(ETFHolding.weight.desc()).all()

    if not holdings:
        raise HTTPException(
            status_code=400,
            detail=f"ETF {normalized_etf} 在 {latest_date} 没有持仓记录，无法校验覆盖范围 {coverage}"
        )

    if coverage_type == "all":
        selected = holdings
    elif coverage_type == "top":
        selected = holdings[:coverage_value]
    else:
        selected = []
        total_weight = 0.0
        for holding in holdings:
            selected.append(holding)
            total_weight += float(holding.weight or 0.0)
            if total_weight >= coverage_value:
                break

    symbols: List[str] = []
    seen = set()
    for holding in selected:
        ticker = _normalize_symbol(getattr(holding, "ticker", None))
        if not ticker or ticker in seen:
            continue
        seen.add(ticker)
        symbols.append(ticker)

    if not symbols:
        raise HTTPException(
            status_code=400,
            detail=f"覆盖范围 {coverage} 未匹配到任何标的，无法导入"
        )
    return symbols


def _extract_import_symbols(
    rows: List[Dict[str, Any]],
    symbol_keys: List[str]
) -> List[str]:
    ordered_keys: List[str] = []
    for key in symbol_keys + ["Ticker", "ticker", "Symbol", "symbol"]:
        if key not in ordered_keys:
            ordered_keys.append(key)

    symbols: List[str] = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        symbol_value = None
        for key in ordered_keys:
            if key in item:
                symbol_value = item.get(key)
                break
        symbol = _normalize_symbol(symbol_value)
        if symbol:
            symbols.append(symbol)
    return symbols


def _format_symbol_list(symbols: List[str], max_items: int = 12) -> str:
    if len(symbols) <= max_items:
        return ", ".join(symbols)
    return f"{', '.join(symbols[:max_items])} 等{len(symbols)}只"


def _assert_import_symbols_match(
    source_label: str,
    etf_symbol: str,
    coverage: str,
    expected_symbols: List[str],
    imported_symbols: List[str],
    required_field_hint: str
) -> Tuple[List[str], List[str]]:
    if not imported_symbols:
        raise HTTPException(
            status_code=400,
            detail=f"{source_label} 导入数据未识别到标的代码，请确认包含字段: {required_field_hint}"
        )

    expected = list(dict.fromkeys(expected_symbols))
    expected_by_key: Dict[str, str] = {}
    for symbol in expected:
        key = _canonical_symbol_key(symbol)
        if key and key not in expected_by_key:
            expected_by_key[key] = symbol

    resolved_imported: List[str] = []
    alias_mappings: List[str] = []
    for symbol in imported_symbols:
        key = _canonical_symbol_key(symbol)
        resolved = expected_by_key.get(key, key)
        resolved_imported.append(resolved)
        raw_normalized = str(symbol).strip().upper()
        if raw_normalized and resolved and raw_normalized != resolved:
            alias_mappings.append(f"{raw_normalized} -> {resolved}")

    imported_counts: Dict[str, int] = {}
    for symbol in resolved_imported:
        imported_counts[symbol] = imported_counts.get(symbol, 0) + 1
    duplicate_symbols = sorted([symbol for symbol, count in imported_counts.items() if count > 1])
    if duplicate_symbols:
        raise HTTPException(
            status_code=400,
            detail=f"{source_label} 导入数据存在重复标的: {', '.join(duplicate_symbols)}"
        )

    imported = list(dict.fromkeys(resolved_imported))
    expected_set = set(expected)
    imported_set = set(imported)
    missing = [symbol for symbol in expected if symbol not in imported_set]
    extra = [symbol for symbol in imported if symbol not in expected_set]

    if len(imported) != len(expected) or missing or extra:
        mismatch_details: List[str] = []
        if len(imported) != len(expected):
            mismatch_details.append(f"数量不一致（期望 {len(expected)}，实际 {len(imported)}）")
        if missing:
            mismatch_details.append(f"缺少: {_format_symbol_list(missing)}")
        if extra:
            mismatch_details.append(f"多出: {_format_symbol_list(extra)}")
        raise HTTPException(
            status_code=400,
            detail=(
                f"{source_label} 覆盖范围校验失败（{etf_symbol} / {coverage}）："
                + "；".join(mismatch_details)
            )
        )

    deduped_alias_mappings = list(dict.fromkeys(alias_mappings))
    return imported, deduped_alias_mappings


def _validate_import_rows_by_coverage(
    db: Session,
    source_label: str,
    etf_symbol: str,
    coverage: str,
    rows: List[Dict[str, Any]],
    symbol_keys: List[str],
    required_field_hint: str
) -> Tuple[List[str], List[str]]:
    expected_symbols = _pick_coverage_symbols(db, etf_symbol, coverage)
    imported_symbols = _extract_import_symbols(rows, symbol_keys)
    return _assert_import_symbols_match(
        source_label=source_label,
        etf_symbol=etf_symbol,
        coverage=coverage,
        expected_symbols=expected_symbols,
        imported_symbols=imported_symbols,
        required_field_hint=required_field_hint
    )

def parse_xlsx_holdings(file_content: bytes) -> List[Dict[str, Any]]:
    """
    解析 xlsx 文件，提取 Ticker 和 Weight 列
    
    支持的表头格式:
    | Name | Ticker | Identifier | SEDOL | Weight | Sector | Shares Held | Local Currency |
    """
    try:
        import openpyxl
        from io import BytesIO
        import re
        
        workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=False, data_only=True)
        
        def normalize_header(value) -> str:
            if value is None:
                return ""
            text = str(value).strip().lower().replace("\u00a0", " ")
            text = re.sub(r"\s+", " ", text)
            return re.sub(r"[^a-z0-9]", "", text)
        
        def find_header_in_sheet(sheet, max_rows: int = 50):
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=max_rows, max_col=sheet.max_column, values_only=True), start=1
            ):
                row_ticker = None
                row_weight = None
                for idx, cell in enumerate(row):
                    header_key = normalize_header(cell)
                    if not header_key:
                        continue
                    if row_ticker is None and ("ticker" in header_key or header_key == "symbol"):
                        row_ticker = idx
                    if row_weight is None and "weight" in header_key:
                        row_weight = idx
                if row_ticker is not None and row_weight is not None:
                    return row_idx, row_ticker, row_weight
            return None, None, None

        header_row = None
        ticker_idx = None
        weight_idx = None
        target_sheet = None
        for sheet in workbook.worksheets:
            header_row, ticker_idx, weight_idx = find_header_in_sheet(sheet)
            if header_row is not None:
                target_sheet = sheet
                break

        if header_row is None or target_sheet is None:
            raise ValueError("未找到包含 Ticker 和 Weight 的表头行")
        
        # 解析数据行
        holdings = []
        for row_idx, row in enumerate(
            target_sheet.iter_rows(
                min_row=header_row + 1,
                max_row=target_sheet.max_row,
                max_col=target_sheet.max_column,
                values_only=True
            ),
            start=header_row + 1
        ):
            if len(row) > max(ticker_idx, weight_idx):
                ticker = row[ticker_idx]
                weight = row[weight_idx]
                
                if ticker and weight is not None:
                    holdings.append({
                        "row": row_idx,
                        "ticker": str(ticker).strip(),
                        "weight": weight
                    })
        
        return holdings
        
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="需要安装 openpyxl 库: pip install openpyxl"
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析 xlsx 文件失败: {str(e)}")


def validate_and_filter_holdings(holdings: List[Dict[str, Any]]) -> tuple:
    """
    验证并过滤持仓数据
    返回: (有效持仓列表, 跳过的记录详情)
    """
    def normalize_weight(value: Any) -> Any:
        if isinstance(value, str):
            cleaned = value.strip().replace(",", "")
            if cleaned.endswith("%"):
                cleaned = cleaned[:-1].strip()
            return cleaned
        return value

    valid_holdings = []
    skipped = []
    
    for h in holdings:
        ticker = h.get("ticker", "")
        weight = h.get("weight")
        row = h.get("row", "unknown")
        
        # 验证 Ticker
        if not is_valid_ticker(ticker):
            skipped.append({
                "row": str(row),
                "ticker": ticker,
                "reason": "Ticker 为空或不是有效的英文字符"
            })
            continue
        
        # 验证 Weight
        try:
            weight_float = float(normalize_weight(weight))
            if weight_float <= 0:
                skipped.append({
                    "row": str(row),
                    "ticker": ticker,
                    "reason": f"Weight 值无效: {weight}"
                })
                continue
        except (ValueError, TypeError):
            skipped.append({
                "row": str(row),
                "ticker": ticker,
                "reason": f"Weight 无法转换为数字: {weight}"
            })
            continue
        
        valid_holdings.append({
            "ticker": ticker.upper(),
            "weight": weight_float
        })
    
    return valid_holdings, skipped


# ==================== API Endpoints ====================

@router.post("/holdings/xlsx", response_model=HoldingsUploadResponse)
async def upload_holdings_xlsx(
    file: UploadFile = File(..., description="xlsx 文件"),
    etf_type: str = Form(..., description="ETF 类型: sector 或 industry"),
    etf_symbol: str = Form(..., description="ETF 符号"),
    data_date: str = Form(..., description="数据日期 (YYYY-MM-DD)"),
    parent_sector: Optional[str] = Form(None, description="父板块符号（仅 industry 类型需要）"),
    db: Session = Depends(get_db)
):
    """
    上传 ETF Holdings xlsx 文件
    
    - 板块 ETF: etf_type=sector, etf_symbol 必须是 11 个默认板块之一
    - 行业 ETF: etf_type=industry, 需要提供 parent_sector（所属板块）
    
    xlsx 文件格式:
    | Name | Ticker | Identifier | SEDOL | Weight | Sector | Shares Held | Local Currency |
    
    只会提取 Ticker 和 Weight 列
    """
    # 验证 etf_type
    if etf_type not in ["sector", "industry"]:
        raise HTTPException(status_code=400, detail="etf_type 必须是 'sector' 或 'industry'")
    
    etf_symbol = etf_symbol.upper()
    
    # 板块 ETF 验证
    if etf_type == "sector":
        if not is_valid_sector_symbol(etf_symbol):
            raise HTTPException(
                status_code=400, 
                detail=f"无效的板块 ETF 符号。有效值: {', '.join(VALID_SECTOR_SYMBOLS)}"
            )
    
    # 行业 ETF 验证
    if etf_type == "industry":
        if parent_sector:
            parent_sector = parent_sector.upper()
            if not is_valid_sector_symbol(parent_sector):
                raise HTTPException(
                    status_code=400,
                    detail=f"无效的父板块符号。有效值: {', '.join(VALID_SECTOR_SYMBOLS)}"
                )
    
    # 验证日期格式
    try:
        parsed_date = datetime.strptime(data_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式无效，请使用 YYYY-MM-DD 格式")
    
    # 验证文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 xlsx 或 xls 文件格式")
    
    try:
        # 读取文件内容
        file_content = await file.read()
        
        # 解析 xlsx
        raw_holdings = parse_xlsx_holdings(file_content)
        
        if not raw_holdings:
            raise HTTPException(status_code=400, detail="xlsx 文件中没有找到有效的持仓数据")
        
        # 验证和过滤数据
        valid_holdings, skipped = validate_and_filter_holdings(raw_holdings)
        
        if not valid_holdings:
            raise HTTPException(
                status_code=400, 
                detail=f"所有持仓数据都无效。跳过 {len(skipped)} 条记录"
            )
        
        # 查找或创建 ETF
        etf = db.query(ETF).filter(ETF.symbol == etf_symbol).first()
        
        if not etf:
            # 创建新的 ETF 记录
            etf = ETF(
                symbol=etf_symbol,
                name=etf_symbol,
                type=etf_type,
                parent_sector=parent_sector if etf_type == "industry" else None,
                score=0.0,
                rank=0,
                delta={"delta3d": None, "delta5d": None},
                completeness=0.0,
                holdings_count=0
            )
            db.add(etf)
            db.flush()
            logger.info(f"创建新的 ETF 记录: {etf_symbol}")
        
        # 删除该 ETF 在指定日期的旧持仓数据
        db.query(ETFHolding).filter(
            ETFHolding.etf_id == etf.id,
            ETFHolding.data_date == parsed_date
        ).delete()
        
        # 插入新的持仓数据
        for h in valid_holdings:
            holding = ETFHolding(
                etf_id=etf.id,
                etf_symbol=etf_symbol,
                ticker=h["ticker"],
                weight=h["weight"],
                data_date=parsed_date
            )
            db.add(holding)
        
        # 更新 ETF 的持仓数量
        etf.holdings_count = len(valid_holdings)
        etf.updated_at = datetime.utcnow()
        
        # 记录上传日志
        upload_log = HoldingsUploadLog(
            etf_symbol=etf_symbol,
            etf_type=etf_type,
            data_date=parsed_date,
            file_name=file.filename,
            records_count=len(valid_holdings),
            skipped_count=len(skipped),
            status="success"
        )
        db.add(upload_log)
        
        db.commit()
        
        logger.info(f"成功上传 {etf_symbol} 的持仓数据: {len(valid_holdings)} 条记录, 跳过 {len(skipped)} 条")
        
        return HoldingsUploadResponse(
            status="success",
            etf_symbol=etf_symbol,
            etf_type=etf_type,
            data_date=data_date,
            records_imported=len(valid_holdings),
            records_skipped=len(skipped),
            skipped_details=skipped[:20] if skipped else None  # 最多返回前 20 条跳过记录
        )
        
    except HTTPException:
        raise
    except Exception as e:
        # 记录失败日志
        try:
            upload_log = HoldingsUploadLog(
                etf_symbol=etf_symbol,
                etf_type=etf_type,
                data_date=parsed_date,
                file_name=file.filename if file else None,
                records_count=0,
                skipped_count=0,
                status="error",
                error_message=str(e)
            )
            db.add(upload_log)
            db.commit()
        except:
            db.rollback()
        
        logger.error(f"上传 Holdings 失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/holdings/logs", response_model=List[dict])
async def get_holdings_upload_logs(
    etf_symbol: Optional[str] = Query(None, description="ETF 符号"),
    limit: int = Query(50, description="返回数量限制"),
    db: Session = Depends(get_db)
):
    """
    获取 Holdings 上传日志
    """
    query = db.query(HoldingsUploadLog)
    
    if etf_symbol:
        query = query.filter(HoldingsUploadLog.etf_symbol == etf_symbol.upper())
    
    logs = query.order_by(HoldingsUploadLog.created_at.desc()).limit(limit).all()
    
    return [
        {
            "id": log.id,
            "etfSymbol": log.etf_symbol,
            "etfType": log.etf_type,
            "dataDate": log.data_date.isoformat() if log.data_date else None,
            "fileName": log.file_name,
            "recordsCount": log.records_count,
            "skippedCount": log.skipped_count,
            "status": log.status,
            "errorMessage": log.error_message,
            "createdAt": log.created_at.isoformat() if log.created_at else None
        }
        for log in logs
    ]


@router.post("/finviz", response_model=FinvizImportResponse)
async def import_finviz_data(
    request: FinvizImportRequest,
    db: Session = Depends(get_db)
):
    """
    导入 Finviz 技术指标数据
    
    数据格式示例:
    ```json
    {
        "etf_symbol": "XLK",
        "coverage": "top20",
        "data": [
            {
                "Ticker": "AAPL",
                "Price": 185.5,
                "Change": 1.23,
                "Volume": 50000000,
                "SMA20": 182.0,
                "SMA50": 178.0,
                "SMA200": 172.0,
                "RSI": 55.5,
                "52W High": 199.0,
                "52W Low": 164.0
            }
        ]
    }
    ```
    
    返回:
    - 导入记录数
    - 广度指标 (% above SMA20/50/200)
    - 数据验证结果
    """
    try:
        from app.services.orchestrator import get_orchestrator

        normalized_etf_symbol = str(request.etf_symbol or "").strip().upper()
        normalized_coverage = str(request.coverage or "top20").strip().lower()
        _, alias_mappings = _validate_import_rows_by_coverage(
            db=db,
            source_label="Finviz",
            etf_symbol=normalized_etf_symbol,
            coverage=normalized_coverage,
            rows=request.data,
            symbol_keys=["Ticker", "ticker", "Symbol", "symbol"],
            required_field_hint="Ticker"
        )
        if alias_mappings:
            logger.info(
                "finviz_symbol_alias_mapped etf=%s coverage=%s mappings=%s",
                normalized_etf_symbol,
                normalized_coverage,
                alias_mappings,
            )

        orchestrator = get_orchestrator()

        # 处理导入
        result = await orchestrator.process_finviz_import(
            etf_symbol=normalized_etf_symbol,
            data=request.data,
            coverage=normalized_coverage
        )
        
        if 'error' in result:
            raise HTTPException(status_code=400, detail=result['error'])
        
        # 保存到数据库（供动能股池计算使用）
        parsed_items = result.get('parsed_data') or []
        try:
            inserted = _upsert_imported_data(db, 'finviz', parsed_items)
            if inserted:
                db.commit()
        except Exception as exc:
            db.rollback()
            logger.warning(f"保存 Finviz 数据失败: {exc}")

        return {
            'status': 'success',
            'etf_symbol': normalized_etf_symbol,
            'coverage': normalized_coverage,
            'records_imported': result.get('records_count', 0),
            'breadth_metrics': result.get('breadth_metrics', {}),
            'validation': result.get('validation', {}),
            'statistics': result.get('statistics'),
            'symbol_alias_mappings': alias_mappings or None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Finviz 数据导入失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/marketchameleon", response_model=MCImportResponse)
async def import_mc_data(
    request: MCImportRequest,
    db: Session = Depends(get_db)
):
    """
    导入 MarketChameleon 期权数据
    
    数据格式示例:
    ```json
    {
        "data": [
            {
                "symbol": "AAPL",
                "ivr": 45.5,
                "iv_hv_ratio": 1.2,
                "rel_notional": 85.0,
                "rel_vol": 1.8,
                "trade_count": 5000,
                "iv30": 25.5,
                "iv60": 26.0,
                "iv90": 27.0,
                "pct_multi_leg": 30.0,
                "pct_contingent": 15.0
            }
        ]
    }
    ```
    
    返回:
    - 导入记录数
    - 热度类型分布
    - 处理后的数据（包含 HeatScore、RiskScore）
    """
    try:
        from app.services.orchestrator import get_orchestrator

        normalized_etf_symbol = str(request.etf_symbol or "").strip().upper()
        normalized_coverage = str(request.coverage or "").strip().lower()
        if normalized_etf_symbol or normalized_coverage:
            if not normalized_etf_symbol or not normalized_coverage:
                raise HTTPException(
                    status_code=400,
                    detail="MarketChameleon 导入校验需要同时提供 etf_symbol 和 coverage"
                )
            _, alias_mappings = _validate_import_rows_by_coverage(
                db=db,
                source_label="MarketChameleon",
                etf_symbol=normalized_etf_symbol,
                coverage=normalized_coverage,
                rows=request.data,
                symbol_keys=["symbol", "Symbol", "Ticker", "ticker"],
                required_field_hint="symbol"
            )
            if alias_mappings:
                logger.info(
                    "mc_symbol_alias_mapped etf=%s coverage=%s mappings=%s",
                    normalized_etf_symbol,
                    normalized_coverage,
                    alias_mappings,
                )

        orchestrator = get_orchestrator()

        # 处理导入
        result = await orchestrator.process_mc_import(request.data)
        
        if 'error' in result:
            raise HTTPException(status_code=400, detail=result['error'])
        
        processed_items = result.get('processed_data') or []
        try:
            inserted = _upsert_imported_data(db, 'marketchameleon', processed_items)
            if inserted:
                db.commit()
        except Exception as exc:
            db.rollback()
            logger.warning(f"保存 MarketChameleon 数据失败: {exc}")

        return {
            'status': 'success',
            'records_imported': result.get('records_count', 0),
            'heat_distribution': result.get('heat_distribution', {}),
            'data': result.get('processed_data')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"MarketChameleon 数据导入失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/finviz/csv")
async def import_finviz_csv(
    file: UploadFile = File(...),
    etf_symbol: str = Form(...),
    coverage: str = Form("top20"),
    db: Session = Depends(get_db)
):
    """
    通过 CSV 文件导入 Finviz 数据
    
    CSV 列要求:
    - Ticker: 股票代码
    - Price: 当前价格
    - Change: 涨跌幅
    - Volume: 成交量
    - SMA20, SMA50, SMA200: 均线
    - RSI: RSI 指标
    - 52W High, 52W Low: 52 周高低
    """
    try:
        # 读取 CSV
        content = await file.read()
        text = content.decode('utf-8')
        
        reader = csv.DictReader(io.StringIO(text))
        data = list(reader)
        
        # 转换数值字段
        for row in data:
            for key in ['Price', 'Change', 'Volume', 'SMA20', 'SMA50', 'SMA200', 
                        'RSI', '52W High', '52W Low', 'Rel Volume']:
                if key in row:
                    try:
                        value = row[key].replace(',', '').replace('%', '')
                        row[key] = float(value) if value else None
                    except (ValueError, AttributeError):
                        row[key] = None

        normalized_etf_symbol = str(etf_symbol or "").strip().upper()
        normalized_coverage = str(coverage or "top20").strip().lower()
        _, alias_mappings = _validate_import_rows_by_coverage(
            db=db,
            source_label="Finviz",
            etf_symbol=normalized_etf_symbol,
            coverage=normalized_coverage,
            rows=data,
            symbol_keys=["Ticker", "ticker", "Symbol", "symbol"],
            required_field_hint="Ticker"
        )
        if alias_mappings:
            logger.info(
                "finviz_csv_symbol_alias_mapped etf=%s coverage=%s mappings=%s",
                normalized_etf_symbol,
                normalized_coverage,
                alias_mappings,
            )

        # 处理导入
        from app.services.orchestrator import get_orchestrator

        orchestrator = get_orchestrator()
        result = await orchestrator.process_finviz_import(
            etf_symbol=normalized_etf_symbol,
            data=data,
            coverage=normalized_coverage
        )
        
        parsed_items = result.get('parsed_data') or []
        try:
            inserted = _upsert_imported_data(db, 'finviz', parsed_items)
            if inserted:
                db.commit()
        except Exception as exc:
            db.rollback()
            logger.warning(f"保存 Finviz CSV 数据失败: {exc}")

        return {
            'status': 'success',
            'etf_symbol': normalized_etf_symbol,
            'coverage': normalized_coverage,
            'records_imported': result.get('records_count', 0),
            'breadth_metrics': result.get('breadth_metrics', {}),
            'validation': result.get('validation', {}),
            'symbol_alias_mappings': alias_mappings or None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"CSV 导入失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/marketchameleon/csv")
async def import_mc_csv(
    file: UploadFile = File(...),
    etf_symbol: Optional[str] = Form(None),
    coverage: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    通过 CSV 文件导入 MarketChameleon 数据
    
    CSV 列要求:
    - symbol: 股票代码
    - ivr: IV Rank
    - iv_hv_ratio: IV/HV 比率
    - rel_notional: 相对名义成交额
    - rel_vol: 相对成交量
    - trade_count: 交易笔数
    - iv30, iv60, iv90: IV 期限结构
    """
    try:
        content = await file.read()
        text = content.decode('utf-8')
        
        reader = csv.DictReader(io.StringIO(text))
        data = list(reader)
        
        # 转换数值字段
        numeric_fields = ['ivr', 'iv_hv_ratio', 'rel_notional', 'rel_vol', 
                          'trade_count', 'iv30', 'iv60', 'iv90', 
                          'pct_multi_leg', 'pct_contingent', 'iv_change']
        
        for row in data:
            for key in numeric_fields:
                if key in row:
                    try:
                        value = row[key].replace(',', '').replace('%', '')
                        row[key] = float(value) if value else None
                    except (ValueError, AttributeError):
                        row[key] = None

        normalized_etf_symbol = str(etf_symbol or "").strip().upper()
        normalized_coverage = str(coverage or "").strip().lower()
        if normalized_etf_symbol or normalized_coverage:
            if not normalized_etf_symbol or not normalized_coverage:
                raise HTTPException(
                    status_code=400,
                    detail="MarketChameleon CSV 导入校验需要同时提供 etf_symbol 和 coverage"
                )
            _, alias_mappings = _validate_import_rows_by_coverage(
                db=db,
                source_label="MarketChameleon",
                etf_symbol=normalized_etf_symbol,
                coverage=normalized_coverage,
                rows=data,
                symbol_keys=["symbol", "Symbol", "Ticker", "ticker"],
                required_field_hint="symbol"
            )
            if alias_mappings:
                logger.info(
                    "mc_csv_symbol_alias_mapped etf=%s coverage=%s mappings=%s",
                    normalized_etf_symbol,
                    normalized_coverage,
                    alias_mappings,
                )

        # 处理导入
        from app.services.orchestrator import get_orchestrator
        
        orchestrator = get_orchestrator()
        result = await orchestrator.process_mc_import(data)
        
        processed_items = result.get('processed_data') or []
        try:
            inserted = _upsert_imported_data(db, 'marketchameleon', processed_items)
            if inserted:
                db.commit()
        except Exception as exc:
            db.rollback()
            logger.warning(f"保存 MarketChameleon CSV 数据失败: {exc}")

        return {
            'status': 'success',
            'records_imported': result.get('records_count', 0),
            'heat_distribution': result.get('heat_distribution', {}),
            'data': result.get('processed_data')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"MarketChameleon CSV 导入失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/holdings", response_model=HoldingsImportResponse)
async def import_etf_holdings(request: HoldingsImportRequest):
    """
    导入 ETF 持仓数据
    
    数据格式:
    ```json
    {
        "etf_symbol": "XLK",
        "holdings": [
            {"ticker": "AAPL", "weight": 22.5},
            {"ticker": "MSFT", "weight": 21.0},
            {"ticker": "NVDA", "weight": 6.5}
        ]
    }
    ```
    
    返回:
    - 持仓数量
    - 持仓评分（如果 IBKR 已连接）
    """
    try:
        holdings_count = len(request.holdings)
        
        # 提取股票代码
        symbols = [h.get('ticker') or h.get('symbol') for h in request.holdings]
        symbols = [s for s in symbols if s]
        
        scored_holdings = None
        
        # 尝试评分持仓
        try:
            from app.services.orchestrator import get_orchestrator
            
            orchestrator = get_orchestrator()
            broker_status = orchestrator.get_broker_status()
            
            if broker_status.get('ibkr', {}).get('is_connected', False):
                scored_holdings = await orchestrator.score_etf_holdings(
                    etf_symbol=request.etf_symbol,
                    holdings=symbols,
                    top_n=20
                )
        except Exception as e:
            logger.warning(f"持仓评分失败: {e}")
        
        return {
            'status': 'success',
            'etf_symbol': request.etf_symbol,
            'holdings_count': holdings_count,
            'scored_holdings': scored_holdings
        }
        
    except Exception as e:
        logger.error(f"ETF 持仓导入失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/finviz")
async def get_finviz_template():
    """
    获取 Finviz 数据导入模板
    """
    template = {
        'description': 'Finviz 数据导入模板',
        'required_fields': ['Ticker'],
        'recommended_fields': [
            'Price', 'Change', 'Volume', 
            'SMA20', 'SMA50', 'SMA200',
            'RSI', '52W High', '52W Low', 'Rel Volume'
        ],
        'sample_data': [
            {
                'Ticker': 'AAPL',
                'Price': 185.50,
                'Change': 1.23,
                'Volume': 50000000,
                'SMA20': 182.0,
                'SMA50': 178.0,
                'SMA200': 172.0,
                'RSI': 55.5,
                '52W High': 199.0,
                '52W Low': 164.0,
                'Rel Volume': 1.2
            }
        ],
        'notes': [
            '从 Finviz Screener 导出数据',
            'Ticker 字段必填',
            '价格和均线字段用于计算广度指标'
        ]
    }
    return template


@router.get("/templates/marketchameleon")
async def get_mc_template():
    """
    获取 MarketChameleon 数据导入模板
    """
    template = {
        'description': 'MarketChameleon 期权数据导入模板',
        'required_fields': ['symbol'],
        'recommended_fields': [
            'ivr', 'iv_hv_ratio',
            'rel_notional', 'rel_vol', 'trade_count',
            'iv30', 'iv60', 'iv90',
            'pct_multi_leg', 'pct_contingent', 'iv_change'
        ],
        'sample_data': [
            {
                'symbol': 'AAPL',
                'ivr': 45.5,
                'iv_hv_ratio': 1.2,
                'rel_notional': 85.0,
                'rel_vol': 1.8,
                'trade_count': 5000,
                'iv30': 25.5,
                'iv60': 26.0,
                'iv90': 27.0,
                'pct_multi_leg': 30.0,
                'pct_contingent': 15.0,
                'iv_change': 2.5
            }
        ],
        'calculated_fields': [
            'heat_score: 基于 rel_notional, rel_vol, trade_count 计算',
            'risk_score: 基于 ivr, iv_hv_ratio, iv_change 计算',
            'confidence_penalty: 基于 pct_multi_leg, pct_contingent 计算',
            'term_score: 基于 iv30, iv60, iv90 斜率计算',
            'heat_type: 热度类型分类'
        ]
    }
    return template


@router.get("/templates/holdings")
async def get_holdings_template():
    """
    获取 ETF Holdings xlsx 导入模板
    """
    template = {
        'description': 'ETF Holdings 数据导入模板',
        'file_format': 'xlsx',
        'required_columns': ['Ticker', 'Weight'],
        'all_columns': [
            'Name', 'Ticker', 'Identifier', 'SEDOL', 
            'Weight', 'Sector', 'Shares Held', 'Local Currency'
        ],
        'sample_data': [
            {'Name': 'Apple Inc.', 'Ticker': 'AAPL', 'Weight': 22.5},
            {'Name': 'Microsoft Corp', 'Ticker': 'MSFT', 'Weight': 21.0},
            {'Name': 'NVIDIA Corp', 'Ticker': 'NVDA', 'Weight': 6.5}
        ],
        'validation_rules': [
            'Ticker 必须是有效的英文字符（以字母开头，可包含数字、点号、短横线）',
            'Ticker 为空或无效时，该行会被忽略',
            'Weight 必须是正数'
        ],
        'valid_sector_etfs': VALID_SECTOR_SYMBOLS,
        'upload_commands': {
            'sector_etf': 'uploads -d YYYY-MM-DD -t sector -a ETF_SYMBOL',
            'industry_etf': 'uploads -d YYYY-MM-DD -t industry -s PARENT_SECTOR -a ETF_SYMBOL'
        }
    }
    return template


@router.delete("/cache")
async def clear_import_cache():
    """
    清除导入数据缓存
    """
    try:
        from app.services.orchestrator import get_orchestrator
        
        orchestrator = get_orchestrator()
        orchestrator.clear_cache()
        
        return {
            'status': 'success',
            'message': 'Import cache cleared'
        }
        
    except Exception as e:
        logger.error(f"清除缓存失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))
