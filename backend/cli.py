#!/usr/bin/env python3
"""
Momentum Radar 命令行工具

支持的命令:
- uploads: 上传 ETF Holdings 文件（支持 xlsx、xls、csv 格式）
- update: 更新 ETF Holdings 数据（与 uploads 相同，更语义化的命令）
- init: 初始化数据库和默认数据

使用示例:
    # 上传板块 ETF holdings
    python cli.py uploads -d 2026-01-25 -t sector -a XLK holdings.xlsx
    
    # 上传行业 ETF holdings（需要指定父板块）
    python cli.py uploads -d 2026-01-25 -t industry -s XLK -a SOXX holdings.csv
    
    # 更新数据（日期可选，默认为当天）
    python cli.py update -t sector -a XLE xle.xlsx
    python cli.py update -d 2026-01-28 -t sector -a XLE xle.xlsx
    
    # 初始化数据库
    python cli.py init
"""

import argparse
import sys
import os
from datetime import datetime
from pathlib import Path

# 添加项目根目录到 Python 路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def parse_xlsx_holdings(file_path: str) -> list:
    """解析 xlsx 文件，提取 Ticker 和 Weight 列"""
    try:
        import openpyxl
        
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # 在 read_only 模式下需要重置维度以正确读取所有列
        sheet.reset_dimensions()
        
        # 获取表头 - 使用 iter_rows 确保在 read_only 模式下正确读取所有列
        headers = []
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell).strip() if cell else "" for cell in row]
            break
        
        # 查找 Ticker 和 Weight 列索引
        ticker_idx = None
        weight_idx = None
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if header_lower == "ticker":
                ticker_idx = idx
            elif header_lower == "weight":
                weight_idx = idx
        
        if ticker_idx is None:
            raise ValueError("未找到 'Ticker' 列")
        if weight_idx is None:
            raise ValueError("未找到 'Weight' 列")
        
        # 解析数据行
        holdings = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) > max(ticker_idx, weight_idx):
                ticker = row[ticker_idx]
                weight = row[weight_idx]
                
                if ticker and weight is not None:
                    holdings.append({
                        "row": row_idx,
                        "ticker": str(ticker).strip(),
                        "weight": weight
                    })
        
        return holdings
        
    except ImportError:
        print("错误: 需要安装 openpyxl 库")
        print("请运行: pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"错误: 解析 xlsx 文件失败 - {e}")
        sys.exit(1)


def parse_csv_holdings(file_path: str) -> list:
    """解析 csv 文件，提取 Ticker 和 Weight 列"""
    try:
        import csv
        
        holdings = []
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            # 查找 Ticker 和 Weight 列（不区分大小写）
            fieldnames_lower = {name.lower(): name for name in reader.fieldnames} if reader.fieldnames else {}
            
            ticker_col = fieldnames_lower.get('ticker')
            weight_col = fieldnames_lower.get('weight')
            
            if not ticker_col:
                raise ValueError("未找到 'Ticker' 列")
            if not weight_col:
                raise ValueError("未找到 'Weight' 列")
            
            for row_idx, row in enumerate(reader, start=2):
                ticker = row.get(ticker_col)
                weight = row.get(weight_col)
                
                if ticker and weight is not None:
                    # 处理可能带有千分位逗号的数字
                    if isinstance(weight, str):
                        weight = weight.replace(',', '')
                    holdings.append({
                        "row": row_idx,
                        "ticker": str(ticker).strip(),
                        "weight": weight
                    })
        
        return holdings
        
    except Exception as e:
        print(f"错误: 解析 csv 文件失败 - {e}")
        sys.exit(1)


def validate_holdings(holdings: list) -> tuple:
    """验证并过滤持仓数据"""
    import re
    
    def is_valid_ticker(ticker: str) -> bool:
        if not ticker or not isinstance(ticker, str):
            return False
        ticker = ticker.strip()
        if not ticker:
            return False
        pattern = r'^[A-Za-z][A-Za-z0-9.\-]*$'
        return bool(re.match(pattern, ticker))
    
    valid_holdings = []
    skipped = []
    
    for h in holdings:
        ticker = h.get("ticker", "")
        weight = h.get("weight")
        row = h.get("row", "unknown")
        
        # 验证 Ticker
        if not is_valid_ticker(ticker):
            skipped.append({
                "row": str(row),
                "ticker": ticker,
                "reason": "Ticker 为空或不是有效的英文字符"
            })
            continue
        
        # 验证 Weight
        try:
            weight_float = float(weight)
            if weight_float <= 0:
                skipped.append({
                    "row": str(row),
                    "ticker": ticker,
                    "reason": f"Weight 值无效: {weight}"
                })
                continue
        except (ValueError, TypeError):
            skipped.append({
                "row": str(row),
                "ticker": ticker,
                "reason": f"Weight 无法转换为数字: {weight}"
            })
            continue
        
        valid_holdings.append({
            "ticker": ticker.upper(),
            "weight": weight_float
        })
    
    return valid_holdings, skipped


def cmd_uploads(args):
    """处理 uploads 命令"""
    from app.models.database import (
        SessionLocal, ETF, ETFHolding, HoldingsUploadLog,
        is_valid_sector_symbol, VALID_SECTOR_SYMBOLS, init_db
    )
    
    # 初始化数据库
    init_db()
    
    # 验证参数
    etf_type = args.type
    etf_symbol = args.etf_symbol.upper()
    data_date_str = args.date if args.date else datetime.now().strftime("%Y-%m-%d")
    file_path = args.file
    parent_sector = args.sector.upper() if args.sector else None
    
    # 验证日期格式
    try:
        data_date = datetime.strptime(data_date_str, "%Y-%m-%d").date()
    except ValueError:
        print(f"错误: 日期格式无效，请使用 YYYY-MM-DD 格式")
        sys.exit(1)
    
    # 验证 ETF 类型
    if etf_type not in ["sector", "industry"]:
        print(f"错误: ETF 类型必须是 'sector' 或 'industry'")
        sys.exit(1)
    
    # 板块 ETF 验证
    if etf_type == "sector":
        if not is_valid_sector_symbol(etf_symbol):
            print(f"错误: 无效的板块 ETF 符号")
            print(f"有效的板块 ETF: {', '.join(VALID_SECTOR_SYMBOLS)}")
            sys.exit(1)
    
    # 行业 ETF 验证
    if etf_type == "industry" and parent_sector:
        if not is_valid_sector_symbol(parent_sector):
            print(f"错误: 无效的父板块符号")
            print(f"有效的板块 ETF: {', '.join(VALID_SECTOR_SYMBOLS)}")
            sys.exit(1)
    
    # 验证文件存在
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}")
        sys.exit(1)
    
    # 验证文件类型
    if not file_path.endswith(('.xlsx', '.xls', '.csv')):
        print(f"错误: 只支持 xlsx、xls 或 csv 文件格式")
        sys.exit(1)
    
    print(f"\n{'='*60}")
    print(f"上传 ETF Holdings")
    print(f"{'='*60}")
    print(f"ETF 类型: {etf_type}")
    print(f"ETF 符号: {etf_symbol}")
    if parent_sector:
        print(f"父板块: {parent_sector}")
    print(f"数据日期: {data_date_str}")
    print(f"文件: {file_path}")
    print(f"{'='*60}\n")
    
    # 解析文件
    if file_path.endswith('.csv'):
        print("正在解析 csv 文件...")
        raw_holdings = parse_csv_holdings(file_path)
    else:
        print("正在解析 xlsx 文件...")
        raw_holdings = parse_xlsx_holdings(file_path)
    print(f"找到 {len(raw_holdings)} 行数据")
    
    # 验证数据
    print("正在验证数据...")
    valid_holdings, skipped = validate_holdings(raw_holdings)
    print(f"有效记录: {len(valid_holdings)} 条")
    print(f"跳过记录: {len(skipped)} 条")
    
    if skipped and len(skipped) <= 10:
        print("\n跳过的记录详情:")
        for s in skipped:
            print(f"  行 {s['row']}: {s['ticker']} - {s['reason']}")
    elif skipped:
        print(f"\n前 10 条跳过记录详情:")
        for s in skipped[:10]:
            print(f"  行 {s['row']}: {s['ticker']} - {s['reason']}")
        print(f"  ... 还有 {len(skipped) - 10} 条")
    
    if not valid_holdings:
        print("\n错误: 没有有效的持仓数据")
        sys.exit(1)
    
    # 写入数据库
    print("\n正在写入数据库...")
    db = SessionLocal()
    
    try:
        # 查找或创建 ETF
        etf = db.query(ETF).filter(ETF.symbol == etf_symbol).first()
        
        if not etf:
            etf = ETF(
                symbol=etf_symbol,
                name=etf_symbol,
                type=etf_type,
                parent_sector=parent_sector if etf_type == "industry" else None,
                score=0.0,
                rank=0,
                delta={"delta3d": None, "delta5d": None},
                completeness=0.0,
                holdings_count=0
            )
            db.add(etf)
            db.flush()
            print(f"创建新的 ETF 记录: {etf_symbol}")
        
        # 删除该 ETF 在指定日期的旧持仓数据
        deleted = db.query(ETFHolding).filter(
            ETFHolding.etf_id == etf.id,
            ETFHolding.data_date == data_date
        ).delete()
        if deleted:
            print(f"删除旧数据: {deleted} 条记录")
        
        # 插入新的持仓数据
        for h in valid_holdings:
            holding = ETFHolding(
                etf_id=etf.id,
                etf_symbol=etf_symbol,
                ticker=h["ticker"],
                weight=h["weight"],
                data_date=data_date
            )
            db.add(holding)
        
        # 更新 ETF 的持仓数量
        etf.holdings_count = len(valid_holdings)
        etf.updated_at = datetime.now()
        
        # 删除该 ETF 在指定日期的旧上传日志（支持重复上传）
        db.query(HoldingsUploadLog).filter(
            HoldingsUploadLog.etf_symbol == etf_symbol,
            HoldingsUploadLog.data_date == data_date
        ).delete()
        
        # 记录上传日志
        upload_log = HoldingsUploadLog(
            etf_symbol=etf_symbol,
            etf_type=etf_type,
            data_date=data_date,
            file_name=os.path.basename(file_path),
            records_count=len(valid_holdings),
            skipped_count=len(skipped),
            status="success"
        )
        db.add(upload_log)
        
        db.commit()
        
        print(f"\n{'='*60}")
        print(f"上传成功!")
        print(f"{'='*60}")
        print(f"ETF: {etf_symbol}")
        print(f"日期: {data_date_str}")
        print(f"导入记录: {len(valid_holdings)} 条")
        print(f"跳过记录: {len(skipped)} 条")
        print(f"{'='*60}\n")
        
    except Exception as e:
        db.rollback()
        print(f"\n错误: 写入数据库失败 - {e}")
        
        # 记录失败日志
        try:
            upload_log = HoldingsUploadLog(
                etf_symbol=etf_symbol,
                etf_type=etf_type,
                data_date=data_date,
                file_name=os.path.basename(file_path),
                records_count=0,
                skipped_count=0,
                status="error",
                error_message=str(e)
            )
            db.add(upload_log)
            db.commit()
        except:
            db.rollback()
        
        sys.exit(1)
    finally:
        db.close()


def cmd_init(args):
    """处理 init 命令"""
    from app.models.database import init_db, init_default_sector_etfs
    
    print("正在初始化数据库...")
    init_db()
    print("数据库表已创建")
    
    print("正在初始化默认板块 ETF...")
    init_default_sector_etfs()
    print("默认板块 ETF 已初始化")
    
    print("\n初始化完成!")


def cmd_list_etfs(args):
    """列出所有 ETF"""
    from app.models.database import SessionLocal, ETF, init_db
    
    init_db()
    db = SessionLocal()
    
    try:
        etf_type = args.type if args.type else None
        
        query = db.query(ETF)
        if etf_type:
            query = query.filter(ETF.type == etf_type)
        
        etfs = query.order_by(ETF.type, ETF.symbol).all()
        
        print(f"\n{'='*70}")
        print(f"ETF 列表 (共 {len(etfs)} 个)")
        print(f"{'='*70}")
        print(f"{'类型':<10} {'符号':<10} {'名称':<30} {'持仓数':<10}")
        print(f"{'-'*70}")
        
        for etf in etfs:
            print(f"{etf.type:<10} {etf.symbol:<10} {etf.name[:28]:<30} {etf.holdings_count:<10}")
        
        print(f"{'='*70}\n")
        
    finally:
        db.close()


def cmd_list_holdings(args):
    """列出 ETF 持仓"""
    from app.models.database import SessionLocal, ETF, ETFHolding, init_db
    from sqlalchemy import func
    
    init_db()
    db = SessionLocal()
    
    try:
        etf_symbol = args.etf_symbol.upper()
        
        etf = db.query(ETF).filter(ETF.symbol == etf_symbol).first()
        if not etf:
            print(f"错误: 未找到 ETF '{etf_symbol}'")
            sys.exit(1)
        
        # 获取最新日期
        if args.date:
            try:
                data_date = datetime.strptime(args.date, "%Y-%m-%d").date()
            except ValueError:
                print(f"错误: 日期格式无效，请使用 YYYY-MM-DD 格式")
                sys.exit(1)
        else:
            data_date = db.query(func.max(ETFHolding.data_date)).filter(
                ETFHolding.etf_id == etf.id
            ).scalar()
        
        if not data_date:
            print(f"错误: ETF '{etf_symbol}' 没有持仓数据")
            sys.exit(1)
        
        holdings = db.query(ETFHolding).filter(
            ETFHolding.etf_id == etf.id,
            ETFHolding.data_date == data_date
        ).order_by(ETFHolding.weight.desc()).all()
        
        print(f"\n{'='*50}")
        print(f"{etf_symbol} 持仓 ({data_date})")
        print(f"{'='*50}")
        print(f"{'#':<5} {'Ticker':<10} {'Weight (%)':<15}")
        print(f"{'-'*50}")
        
        for idx, h in enumerate(holdings, 1):
            print(f"{idx:<5} {h.ticker:<10} {h.weight:<15.2f}")
        
        print(f"{'-'*50}")
        print(f"总计: {len(holdings)} 个持仓")
        print(f"{'='*50}\n")
        
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(
        description='Momentum Radar 命令行工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 上传板块 ETF holdings
  python cli.py uploads -d 2026-01-25 -t sector -a XLK holdings.xlsx
  
  # 上传行业 ETF holdings（需要指定父板块）
  python cli.py uploads -d 2026-01-25 -t industry -s XLK -a SOXX holdings.xlsx
  
  # 更新数据（日期可选，默认为当天）
  python cli.py update -t sector -a XLE xle.xlsx
  python cli.py update -d 2026-01-28 -t sector -a XLE xle.xlsx
  
  # 初始化数据库
  python cli.py init
  
  # 列出所有 ETF
  python cli.py list-etfs
  
  # 列出 ETF 持仓
  python cli.py list-holdings XLK
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='可用命令')
    
    # uploads 命令
    uploads_parser = subparsers.add_parser('uploads', help='上传 ETF Holdings 文件 (xlsx/xls/csv)')
    uploads_parser.add_argument('-d', '--date', required=False, help='数据日期 (YYYY-MM-DD)，默认为当天')
    uploads_parser.add_argument('-t', '--type', required=True, choices=['sector', 'industry'],
                               help='ETF 类型: sector 或 industry')
    uploads_parser.add_argument('-a', '--etf-symbol', required=True, dest='etf_symbol',
                               help='ETF 符号 (如 XLK, SOXX)')
    uploads_parser.add_argument('-s', '--sector', help='父板块符号 (仅 industry 类型需要)')
    uploads_parser.add_argument('file', help='Holdings 文件路径 (xlsx/xls/csv)')
    uploads_parser.set_defaults(func=cmd_uploads)
    
    # update 命令 (uploads 的别名，更语义化)
    update_parser = subparsers.add_parser('update', help='更新 ETF Holdings 数据 (与 uploads 相同)')
    update_parser.add_argument('-d', '--date', required=False, help='数据日期 (YYYY-MM-DD)，默认为当天')
    update_parser.add_argument('-t', '--type', required=True, choices=['sector', 'industry'],
                              help='ETF 类型: sector 或 industry')
    update_parser.add_argument('-a', '--etf-symbol', required=True, dest='etf_symbol',
                              help='ETF 符号 (如 XLK, SOXX)')
    update_parser.add_argument('-s', '--sector', help='父板块符号 (仅 industry 类型需要)')
    update_parser.add_argument('file', help='Holdings 文件路径 (xlsx/xls/csv)')
    update_parser.set_defaults(func=cmd_uploads)
    
    # init 命令
    init_parser = subparsers.add_parser('init', help='初始化数据库和默认数据')
    init_parser.set_defaults(func=cmd_init)
    
    # list-etfs 命令
    list_etfs_parser = subparsers.add_parser('list-etfs', help='列出所有 ETF')
    list_etfs_parser.add_argument('-t', '--type', choices=['sector', 'industry'],
                                  help='筛选 ETF 类型')
    list_etfs_parser.set_defaults(func=cmd_list_etfs)
    
    # list-holdings 命令
    list_holdings_parser = subparsers.add_parser('list-holdings', help='列出 ETF 持仓')
    list_holdings_parser.add_argument('etf_symbol', help='ETF 符号')
    list_holdings_parser.add_argument('-d', '--date', help='数据日期 (默认最新)')
    list_holdings_parser.set_defaults(func=cmd_list_holdings)
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    args.func(args)


if __name__ == '__main__':
    main()